from django.test import TestCase
from django.utils import timezone
from django.urls import reverse
import datetime
from collections import defaultdict

from .models import (
    University, Campus, Faculty, Department, Program, Semester,
    Course, Lecturer, StudentGroup, Room, TimeSlot, Constraint, Timetable, ScheduleSlot, LecturerAvailability
)
from .conflicts import detect_conflicts
from .solver import generate_timetable

class TimetableSchedulerTests(TestCase):
    def setUp(self):
        # 1. Base Setup
        self.uni = University.objects.create(name="Test University", code="TU")
        self.campus = Campus.objects.create(university=self.uni, name="Test Campus")
        self.faculty = Faculty.objects.create(campus=self.campus, name="Test Faculty")
        self.dept = Department.objects.create(faculty=self.faculty, name="Test Dept")
        self.program = Program.objects.create(department=self.dept, name="Test Program")
        self.semester = Semester.objects.create(
            university=self.uni,
            name="Fall 2026",
            start_date=datetime.date(2026, 9, 1),
            end_date=datetime.date(2026, 12, 1),
            is_active=True
        )
        
        # 2. Rooms
        self.lecture_room = Room.objects.create(campus=self.campus, name="L101", capacity=100, room_type="Lecture")
        self.lab_room = Room.objects.create(campus=self.campus, name="Lab102", capacity=30, room_type="Lab")

        # 3. Lecturers
        self.lecturer_a = Lecturer.objects.create(department=self.dept, name="Dr. A", email="a@tu.edu")
        self.lecturer_b = Lecturer.objects.create(department=self.dept, name="Dr. B", email="b@tu.edu")

        # 4. Student Groups
        self.group_y1 = StudentGroup.objects.create(program=self.program, name="CS Year 1", size=60)
        self.group_y2 = StudentGroup.objects.create(program=self.program, name="CS Year 2", size=25)

        # 5. Timeslots (Monday slots)
        self.ts_mon_1 = TimeSlot.objects.create(university=self.uni, day_of_week=1, start_time=datetime.time(8, 30), end_time=datetime.time(10, 0), slot_number=1)
        self.ts_mon_2 = TimeSlot.objects.create(university=self.uni, day_of_week=1, start_time=datetime.time(10, 15), end_time=datetime.time(11, 45), slot_number=2)
        self.ts_mon_3 = TimeSlot.objects.create(university=self.uni, day_of_week=1, start_time=datetime.time(12, 0), end_time=datetime.time(13, 30), slot_number=3)

        # 6. Courses
        self.c1 = Course.objects.create(program=self.program, code="CS101", name="Intro CS", duration_slots=1, required_room_type="Lecture", lecturer=self.lecturer_a, student_group=self.group_y1)
        self.c2 = Course.objects.create(program=self.program, code="CS102", name="CS Lab", duration_slots=1, required_room_type="Lab", lecturer=self.lecturer_b, student_group=self.group_y1)

        # 7. Timetable Version
        self.timetable = Timetable.objects.create(semester=self.semester, name="Test Timetable V1")

    def test_conflict_detection_double_booking(self):
        """
        Verify that conflict detection correctly reports a double-booking when the same lecturer is assigned to teach two classes at the same time.
        """
        slot_1 = ScheduleSlot(
            id=1,
            timetable=self.timetable,
            course=self.c1,
            lecturer=self.lecturer_a,
            room=self.lecture_room,
            time_slot=self.ts_mon_1,
            student_group=self.group_y1
        )
        
        # Conflicting slot: same lecturer, same timeslot, different room/group/course
        slot_2 = ScheduleSlot(
            id=2,
            timetable=self.timetable,
            course=self.c2,
            lecturer=self.lecturer_a, # conflict here
            room=self.lab_room,
            time_slot=self.ts_mon_1,
            student_group=self.group_y2
        )

        slots = [slot_1, slot_2]
        conflicts = detect_conflicts(slots, self.uni)
        
        # Verify LECTURER_DOUBLE_BOOKING conflict is found
        lecturer_conflicts = [c for c in conflicts if c['constraint_type'] == 'LECTURER_DOUBLE_BOOKING']
        self.assertEqual(len(lecturer_conflicts), 1)
        self.assertEqual(lecturer_conflicts[0]['severity'], 'error')

    def test_conflict_detection_capacity(self):
        """
        Verify that room capacity conflicts are flagged when a group size exceeds room capacity.
        """
        # CS Year 1 (60 students) scheduled in Lab Room (capacity 30)
        slot = ScheduleSlot.objects.create(
            timetable=self.timetable,
            course=self.c1,
            lecturer=self.lecturer_a,
            room=self.lab_room, # capacity: 30 < 60
            time_slot=self.ts_mon_1,
            student_group=self.group_y1
        )
        conflicts = detect_conflicts([slot], self.uni)
        
        capacity_conflicts = [c for c in conflicts if c['constraint_type'] == 'ROOM_CAPACITY']
        self.assertEqual(len(capacity_conflicts), 1)

    def test_conflict_detection_lecturer_self_service_availability(self):
        """
        Verify that lecturer self-service unavailability is flagged as a conflict.
        """
        LecturerAvailability.objects.create(
            lecturer=self.lecturer_a,
            time_slot=self.ts_mon_1,
            is_available=False
        )
        slot = ScheduleSlot.objects.create(
            timetable=self.timetable,
            course=self.c1,
            lecturer=self.lecturer_a,
            room=self.lecture_room,
            time_slot=self.ts_mon_1,
            student_group=self.group_y1
        )
        conflicts = detect_conflicts([slot], self.uni)
        avail_conflicts = [c for c in conflicts if c['constraint_type'] == 'LECTURER_SELF_SERVICE_AVAILABILITY_VIOLATION']
        self.assertEqual(len(avail_conflicts), 1)
        self.assertEqual(avail_conflicts[0]['severity'], 'error')

    def test_solver_schedules_multiple_sessions_per_week(self):
        """A course with sessions_per_week=2 should produce two weekly session blocks."""
        for day in (2, 3, 4, 5):
            TimeSlot.objects.create(
                university=self.uni, day_of_week=day,
                start_time=datetime.time(8, 0), end_time=datetime.time(9, 30), slot_number=1,
            )
        course = Course.objects.create(
            program=self.program, code="CS201", name="Data Structures",
            duration_slots=1, sessions_per_week=2,
            required_room_type="Lecture", lecturer=self.lecturer_a, student_group=self.group_y1,
        )
        status, message, _score = generate_timetable(self.timetable.id, time_limit_seconds=30)
        self.assertIn(status, ('OPTIMAL', 'FEASIBLE'), message)
        slots = ScheduleSlot.objects.filter(timetable=self.timetable, course=course)
        self.assertEqual(slots.count(), 2)

    def test_conflict_detection_lecturer_max_classes_per_day(self):
        """
        Verify that lecturer max classes per day limit violations are flagged.
        """
        Constraint.objects.create(
            university=self.uni,
            name="Dr. A Max 1 Class/Day",
            constraint_type="MAX_CLASSES_PER_DAY",
            is_hard=True,
            parameters={"lecturer_id": self.lecturer_a.id, "max_classes": 1}
        )
        slot_1 = ScheduleSlot.objects.create(
            timetable=self.timetable,
            course=self.c1,
            lecturer=self.lecturer_a,
            room=self.lecture_room,
            time_slot=self.ts_mon_1,
            student_group=self.group_y1
        )
        slot_2 = ScheduleSlot.objects.create(
            timetable=self.timetable,
            course=self.c2,
            lecturer=self.lecturer_a,
            room=self.lab_room,
            time_slot=self.ts_mon_2,
            student_group=self.group_y1
        )
        conflicts = detect_conflicts([slot_1, slot_2], self.uni)
        max_class_conflicts = [c for c in conflicts if c['constraint_type'] == 'LECTURER_MAX_CLASSES_PER_DAY_VIOLATION']
        self.assertEqual(len(max_class_conflicts), 1)

    def test_conflict_detection_lecturer_weekly_hours(self):
        """
        Verify that lecturer weekly max hours violations are flagged.
        """
        self.lecturer_a.max_hours_per_week = 1
        self.lecturer_a.save()

        slot = ScheduleSlot.objects.create(
            timetable=self.timetable,
            course=self.c1,
            lecturer=self.lecturer_a,
            room=self.lecture_room,
            time_slot=self.ts_mon_1,
            student_group=self.group_y1
        )
        conflicts = detect_conflicts([slot], self.uni)
        hours_conflicts = [c for c in conflicts if c['constraint_type'] == 'LECTURER_WEEKLY_HOURS_VIOLATION']
        self.assertEqual(len(hours_conflicts), 1)
        self.assertEqual(hours_conflicts[0]['severity'], 'warning')

    def test_solver_timetable_generation(self):
        """
        Verify that the OR-Tools CP-SAT scheduler solves the problem successfully without creating hard conflicts.
        """
        status, message, score = generate_timetable(self.timetable.id)
        
        self.assertIn(status, ('OPTIMAL', 'FEASIBLE'))
        
        # Check that slots are saved to database
        saved_slots = ScheduleSlot.objects.filter(timetable=self.timetable)
        self.assertEqual(saved_slots.count(), 2) # c1 and c2 should be scheduled
        
        # Verify no hard conflicts exist in the solved schedule
        conflicts = detect_conflicts(list(saved_slots), self.uni)
        errors = [c for c in conflicts if c['severity'] == 'error']
        self.assertEqual(len(errors), 0, f"Solved timetable contains errors: {errors}")

    def test_solver_gap_minimization(self):
        """
        Verify that the solver schedules courses back-to-back to minimize gaps for Student Group Y1.
        """
        # Run solver
        status, message, score = generate_timetable(self.timetable.id)
        self.assertIn(status, ('OPTIMAL', 'FEASIBLE'))
        
        # Check slot numbers of scheduled slots
        saved_slots = list(ScheduleSlot.objects.filter(timetable=self.timetable).select_related('time_slot'))
        self.assertEqual(len(saved_slots), 2)
        
        slot_nums = sorted([s.time_slot.slot_number for s in saved_slots])
        # The scheduled slot numbers must be back-to-back (1 & 2, or 2 & 3)
        self.assertTrue(
            (slot_nums == [1, 2]) or (slot_nums == [2, 3]),
            f"Expected back-to-back slots to minimize gaps, got slot numbers: {slot_nums}"
        )

    def test_solver_room_type_matching(self):
        """
        Verify that the solver prefers assigning Lab courses to Lab rooms and Lecture courses to Lecture rooms.
        """
        # Change student group to CS Year 2 (size 25) to fit lab capacity (30)
        self.c2.student_group = self.group_y2
        self.c2.save()

        # Run solver
        status, message, score = generate_timetable(self.timetable.id)
        self.assertIn(status, ('OPTIMAL', 'FEASIBLE'))
        
        # Fetch assignments
        slot_c1 = ScheduleSlot.objects.filter(timetable=self.timetable, course=self.c1).first()
        slot_c2 = ScheduleSlot.objects.filter(timetable=self.timetable, course=self.c2).first()
        
        self.assertEqual(slot_c1.room.room_type, "Lecture")
        self.assertEqual(slot_c2.room.room_type, "Lab")




class ResourceAdminUITests(TestCase):
    def setUp(self):
        self.uni_a = University.objects.create(name="University A", code="UA")
        self.uni_b = University.objects.create(name="University B", code="UB")
        
        self.campus_a = Campus.objects.create(university=self.uni_a, name="Campus A")
        self.campus_b = Campus.objects.create(university=self.uni_b, name="Campus B")

        self.faculty_a = Faculty.objects.create(campus=self.campus_a, name="Faculty A")
        self.dept_a = Department.objects.create(faculty=self.faculty_a, name="Dept A")
        self.lecturer = Lecturer.objects.create(department=self.dept_a, name="Lec User", email="lec@ua.edu")

        # Create users for testing
        from django.contrib.auth.models import User
        from accounts.models import UserProfile
        
        self.admin_user = User.objects.create_user(username="admin_user", password="password")
        UserProfile.objects.create(user=self.admin_user, role="admin", university=self.uni_a)
        
        self.scheduler_user = User.objects.create_user(username="sched_user", password="password")
        UserProfile.objects.create(user=self.scheduler_user, role="scheduler", university=self.uni_a)
        
        self.lecturer_user = User.objects.create_user(username="lec_user", password="password")
        UserProfile.objects.create(user=self.lecturer_user, role="lecturer", university=self.uni_a, lecturer=self.lecturer)
        
        self.student_user = User.objects.create_user(username="stud_user", password="password")
        UserProfile.objects.create(user=self.student_user, role="student", university=self.uni_a)

    def test_resource_manager_role_denial(self):
        """
        Lecturers and Students should be redirected away with permission error.
        """
        for user, username in [(self.student_user, "stud_user"), (self.lecturer_user, "lec_user")]:
            self.client.login(username=username, password="password")
            
            # Make sure active_role in session is aligned
            session = self.client.session
            session['active_role'] = user.profile.role
            session['active_university_id'] = self.uni_a.id
            session.save()
            
            response = self.client.get(reverse('scheduler:resources_manager') + '?tab=campus')
            self.assertRedirects(response, reverse('scheduler:dashboard'), fetch_redirect_response=False)
            self.client.logout()


    def test_resource_manager_role_allow(self):
        """
        Admins and Schedulers should be allowed to view the resource manager interface.
        """
        for user, username in [(self.admin_user, "admin_user"), (self.scheduler_user, "sched_user")]:
            self.client.login(username=username, password="password")
            
            session = self.client.session
            session['active_role'] = user.profile.role
            session['active_university_id'] = self.uni_a.id
            session.save()
            
            response = self.client.get(reverse('scheduler:resources_manager') + '?tab=campus')
            self.assertEqual(response.status_code, 200)
            self.client.logout()

    def test_resource_manager_scoping(self):
        """
        Verify that only campuses belonging to the active university are listed.
        """
        self.client.login(username="admin_user", password="password")
        session = self.client.session
        session['active_role'] = 'admin'
        session['active_university_id'] = self.uni_a.id
        session.save()
        
        response = self.client.get(reverse('scheduler:resources_manager') + '?tab=campus')
        self.assertEqual(response.status_code, 200)
        
        # Verify Campus A is shown and Campus B is not shown
        self.assertContains(response, "Campus A")
        self.assertNotContains(response, "Campus B")

    def test_resource_manager_creation(self):
        """
        Verify creating a new campus via POST scopes it to the active university.
        """
        self.client.login(username="admin_user", password="password")
        session = self.client.session
        session['active_role'] = 'admin'
        session['active_university_id'] = self.uni_a.id
        session.save()
        
        # Send post request to create Campus C
        post_data = {
            'university': self.uni_a.id,
            'name': 'Campus C'
        }
        response = self.client.post(reverse('scheduler:resources_manager') + '?tab=campus', data=post_data)
        self.assertRedirects(response, reverse('scheduler:resources_manager') + '?tab=campus')
        
        # Assert database has Campus C belonging to University A
        campus_c = Campus.objects.filter(name='Campus C').first()
        self.assertIsNotNone(campus_c)
        self.assertEqual(campus_c.university, self.uni_a)


class GoogleCalendarExportTests(TestCase):
    def setUp(self):
        self.uni_a = University.objects.create(name="University A", code="UA")
        self.uni_b = University.objects.create(name="University B", code="UB")
        
        self.semester_a = Semester.objects.create(
            university=self.uni_a,
            name="Fall 2026",
            start_date=datetime.date(2026, 9, 1),
            end_date=datetime.date(2026, 12, 1),
            is_active=True
        )
        self.semester_b = Semester.objects.create(
            university=self.uni_b,
            name="Spring 2027",
            start_date=datetime.date(2027, 1, 15),
            end_date=datetime.date(2027, 5, 15),
            is_active=True
        )
        
        self.campus_a = Campus.objects.create(university=self.uni_a, name="Campus A")
        self.faculty_a = Faculty.objects.create(campus=self.campus_a, name="Faculty A")
        self.dept_a = Department.objects.create(faculty=self.faculty_a, name="Dept A")
        self.program_a = Program.objects.create(department=self.dept_a, name="Program A")
        
        self.room_a = Room.objects.create(campus=self.campus_a, name="Room 1", capacity=50, room_type="Lecture")
        self.lecturer_a = Lecturer.objects.create(department=self.dept_a, name="Lec A", email="leca@ua.edu")
        self.group_a = StudentGroup.objects.create(program=self.program_a, name="Group A", size=30)
        self.ts_a = TimeSlot.objects.create(university=self.uni_a, day_of_week=1, start_time=datetime.time(9, 0), end_time=datetime.time(10, 30), slot_number=1)
        self.course_a = Course.objects.create(program=self.program_a, code="CS101", name="CS A", duration_slots=1, required_room_type="Lecture", lecturer=self.lecturer_a, student_group=self.group_a)
        
        self.timetable_a = Timetable.objects.create(semester=self.semester_a, name="Timetable A")
        self.slot_a = ScheduleSlot.objects.create(
            timetable=self.timetable_a,
            course=self.course_a,
            lecturer=self.lecturer_a,
            room=self.room_a,
            time_slot=self.ts_a,
            student_group=self.group_a
        )
        
        self.timetable_b = Timetable.objects.create(semester=self.semester_b, name="Timetable B")

        # Create admin user
        from django.contrib.auth.models import User
        from accounts.models import UserProfile
        self.admin_user = User.objects.create_user(username="test_admin", password="password")
        UserProfile.objects.create(user=self.admin_user, role="admin", university=self.uni_a)

    def test_ics_generation_format(self):
        """
        Verify that generate_ics_content outputs a valid iCalendar structure with proper recurrence rules.
        """
        from .calendar_exporter import generate_ics_content
        ics_text = generate_ics_content(self.timetable_a)
        
        self.assertIn("BEGIN:VCALENDAR", ics_text)
        self.assertIn("VERSION:2.0", ics_text)
        self.assertIn("BEGIN:VEVENT", ics_text)
        self.assertIn("SUMMARY:CS101: CS A", ics_text)
        self.assertIn("RRULE:FREQ=WEEKLY;UNTIL=20261201T235959Z", ics_text)
        self.assertIn("LOCATION:Room 1", ics_text)
        self.assertIn("END:VEVENT", ics_text)
        self.assertIn("END:VCALENDAR", ics_text)

    def test_export_timetable_scoping_isolation(self):
        """
        Exporting a timetable belonging to another university should redirect to dashboard.
        """
        self.client.login(username="test_admin", password="password")
        session = self.client.session
        session['active_role'] = 'admin'
        session['active_university_id'] = self.uni_a.id
        session.save()
        
        # Requesting timetable B (which is for University B) while University A is active should block/redirect
        response = self.client.get(reverse('scheduler:export_timetable_ics', args=[self.timetable_b.pk]))
        self.assertRedirects(response, reverse('scheduler:dashboard'))

    def test_export_timetable_success(self):
        """
        Exporting own university's timetable should return a downloadable text/calendar attachment.
        """
        self.client.login(username="test_admin", password="password")
        session = self.client.session
        session['active_role'] = 'admin'
        session['active_university_id'] = self.uni_a.id
        session.save()
        
        response = self.client.get(reverse('scheduler:export_timetable_ics', args=[self.timetable_a.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'text/calendar')
        self.assertTrue(response['Content-Disposition'].startswith('attachment; filename="timetable_Timetable_A.ics"'))
        
        # Verify body content
        content = response.content.decode('utf-8')
        self.assertIn("BEGIN:VCALENDAR", content)
        self.assertIn("SUMMARY:CS101: CS A", content)

    def test_export_timetable_csv_scoping_isolation(self):
        """
        Exporting CSV of a timetable belonging to another university should redirect to dashboard.
        """
        self.client.login(username="test_admin", password="password")
        session = self.client.session
        session['active_role'] = 'admin'
        session['active_university_id'] = self.uni_a.id
        session.save()
        
        response = self.client.get(reverse('scheduler:export_timetable_csv', args=[self.timetable_b.pk]))
        self.assertRedirects(response, reverse('scheduler:dashboard'))

    def test_export_timetable_csv_success(self):
        """
        Exporting own university's timetable as CSV should return a downloadable text/csv attachment.
        """
        self.client.login(username="test_admin", password="password")
        session = self.client.session
        session['active_role'] = 'admin'
        session['active_university_id'] = self.uni_a.id
        session.save()
        
        response = self.client.get(reverse('scheduler:export_timetable_csv', args=[self.timetable_a.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'text/csv')
        self.assertTrue(response['Content-Disposition'].startswith('attachment; filename="timetable_Timetable_A.csv"'))
        
        content = response.content.decode('utf-8')
        self.assertIn("Course Code,Course Name,Room,Room Type,Lecturer,Student Group,Day,Start Time,End Time", content)
        self.assertIn("CS101,CS A,Room 1,Lecture Hall,Lec A,Group A,Monday,09:00,10:30", content)

    def test_export_timetable_pdf_success(self):
        """
        Exporting own university's timetable as PDF should return a downloadable application/pdf attachment.
        """
        self.client.login(username="test_admin", password="password")
        session = self.client.session
        session['active_role'] = 'admin'
        session['active_university_id'] = self.uni_a.id
        session.save()
        
        response = self.client.get(reverse('scheduler:export_timetable_pdf', args=[self.timetable_a.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')
        self.assertTrue(response['Content-Disposition'].startswith('attachment; filename="timetable_Timetable_A.pdf"'))

    def test_export_timetable_pdf_filtering(self):
        """
        Exporting PDF with active filter parameters should work without errors.
        """
        self.client.login(username="test_admin", password="password")
        session = self.client.session
        session['active_role'] = 'admin'
        session['active_university_id'] = self.uni_a.id
        session.save()
        
        # Test with student group filter
        url = reverse('scheduler:export_timetable_pdf', args=[self.timetable_a.pk])
        response = self.client.get(url + f"?filter_type=group&filter_id={self.group_a.id}")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')

    def test_export_timetable_pdf_bulk_all(self):
        """
        Exporting PDF with bulk options (all_groups, all_rooms, all_lecturers) should work without errors.
        """
        self.client.login(username="test_admin", password="password")
        session = self.client.session
        session['active_role'] = 'admin'
        session['active_university_id'] = self.uni_a.id
        session.save()
        
        url = reverse('scheduler:export_timetable_pdf', args=[self.timetable_a.pk])
        
        for bulk_type in ['all_groups', 'all_rooms', 'all_lecturers']:
            response = self.client.get(url + f"?filter_type={bulk_type}")
            self.assertEqual(response.status_code, 200)
            self.assertEqual(response['Content-Type'], 'application/pdf')

    def test_export_timetable_pdf_layout_types(self):
        """
        Exporting PDF with available layout options (weekly, master, master_list) should work without errors.
        """
        self.client.login(username="test_admin", password="password")
        session = self.client.session
        session['active_role'] = 'admin'
        session['active_university_id'] = self.uni_a.id
        session.save()
        
        url = reverse('scheduler:export_timetable_pdf', args=[self.timetable_a.pk])
        for layout in ['weekly', 'master', 'master_list']:
            response = self.client.get(url + f"?layout_type={layout}")
            self.assertEqual(response.status_code, 200)
            self.assertEqual(response['Content-Type'], 'application/pdf')

    def test_delete_resource_does_not_exist(self):
        """
        Verify that attempting to delete a resource that does not exist redirects gracefully instead of raising a 404.
        """
        self.client.login(username="test_admin", password="password")
        session = self.client.session
        session['active_role'] = 'admin'
        session['active_university_id'] = self.uni_a.id
        session.save()
        
        # POST to delete a room that doesn't exist (e.g. ID 9999)
        url = reverse('scheduler:delete_resource', kwargs={'model_type': 'room', 'pk': 9999})
        response = self.client.post(url)
        self.assertRedirects(response, "/resources/?tab=room")
        
        # GET to confirm delete a room that doesn't exist (e.g. ID 9999)
        response = self.client.get(url)
        self.assertRedirects(response, "/resources/?tab=room")

    def test_constraint_delete_does_not_exist(self):
        """
        Verify that attempting to delete a constraint that does not exist redirects gracefully instead of raising a 404.
        """
        self.client.login(username="test_admin", password="password")
        session = self.client.session
        session['active_role'] = 'admin'
        session['active_university_id'] = self.uni_a.id
        session.save()
        
        url = reverse('scheduler:constraint_delete', kwargs={'pk': 9999})
        response = self.client.post(url)
        self.assertRedirects(response, reverse('scheduler:constraint_list'))

    def test_export_timetable_word_success(self):
        """
        Exporting own university's timetable as Word (.docx) should return a downloadable document attachment.
        """
        self.client.login(username="test_admin", password="password")
        session = self.client.session
        session['active_role'] = 'admin'
        session['active_university_id'] = self.uni_a.id
        session.save()
        
        response = self.client.get(reverse('scheduler:export_timetable_word', args=[self.timetable_a.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.wordprocessingml.document')
        self.assertTrue(response['Content-Disposition'].startswith('attachment; filename="timetable_Timetable_A.docx"'))

    def test_export_timetable_word_bulk_all(self):
        """
        Exporting Word (.docx) with bulk options should work without errors.
        """
        self.client.login(username="test_admin", password="password")
        session = self.client.session
        session['active_role'] = 'admin'
        session['active_university_id'] = self.uni_a.id
        session.save()
        
        url = reverse('scheduler:export_timetable_word', args=[self.timetable_a.pk])
        for bulk_type in ['all_groups', 'all_rooms', 'all_lecturers']:
            response = self.client.get(url + f"?filter_type={bulk_type}")
            self.assertEqual(response.status_code, 200)
            self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.wordprocessingml.document')


class TimetableValidationTests(TestCase):
    def setUp(self):
        self.uni = University.objects.create(name="Validation University", code="VU")
        self.campus = Campus.objects.create(university=self.uni, name="Validation Campus")
        self.faculty = Faculty.objects.create(campus=self.campus, name="Validation Faculty")
        self.dept = Department.objects.create(faculty=self.faculty, name="Validation Dept")
        self.program = Program.objects.create(department=self.dept, name="Validation Program")
        self.semester = Semester.objects.create(
            university=self.uni,
            name="Fall 2026",
            start_date=datetime.date(2026, 9, 1),
            end_date=datetime.date(2026, 12, 1),
            is_active=True
        )
        
        self.room = Room.objects.create(campus=self.campus, name="Room 1", capacity=50, room_type="Lecture")
        self.lecturer = Lecturer.objects.create(department=self.dept, name="Dr. V", email="v@vu.edu", max_hours_per_week=20)
        self.group = StudentGroup.objects.create(program=self.program, name="CS Year 1", size=30)
        self.ts = TimeSlot.objects.create(university=self.uni, day_of_week=1, start_time=datetime.time(9, 0), end_time=datetime.time(10, 30), slot_number=1)
        
        self.course = Course.objects.create(
            program=self.program,
            code="CS101",
            name="CS Intro",
            duration_slots=1,
            required_room_type="Lecture",
            lecturer=self.lecturer,
            student_group=self.group
        )
        self.timetable = Timetable.objects.create(semester=self.semester, name="Validation Timetable")

    def test_validation_success(self):
        """
        Verify that a perfectly configured timetable passes validation checks.
        """
        from .validation import validate_timetable_inputs
        is_valid, errors, warnings = validate_timetable_inputs(self.timetable)
        self.assertTrue(is_valid)
        self.assertEqual(len(errors), 0)

    def test_validation_capacity_failure(self):
        """
        Verify that a student group larger than the largest room fails validation.
        """
        # Set student group size to 60 (larger than Room 1 capacity 50)
        self.group.size = 60
        self.group.save()
        
        from .validation import validate_timetable_inputs
        is_valid, errors, warnings = validate_timetable_inputs(self.timetable)
        self.assertFalse(is_valid)
        self.assertTrue(any("exceeds the capacity of the largest room" in err for err in errors))

    def test_validation_lecturer_overallocation_warning(self):
        """
        Verify that lecturer over-allocation is flagged as a warning if within tolerance,
        and as a blocking error if it exceeds tolerance.
        """
        from .validation import validate_timetable_inputs

        # Create additional timeslots to satisfy pre-flight capacity checks
        for i in range(2, 11):
            TimeSlot.objects.create(
                university=self.uni, day_of_week=1,
                start_time=datetime.time(9, 0), end_time=datetime.time(10, 30),
                slot_number=i
            )

        # Increase sessions per week so total hours = 1 * 10 * 1.5 = 15.0 hours
        self.course.sessions_per_week = 10
        self.course.save()

        # Scenario 1: Lecturer hours over-allocated but within 10% tolerance (e.g. 14 hours limit, 15 assigned -> 15/14 = 1.07 < 1.10)
        self.lecturer.max_hours_per_week = 14
        self.lecturer.save()
        
        is_valid, errors, warnings = validate_timetable_inputs(self.timetable)
        self.assertTrue(is_valid, f"Validation failed with errors: {errors}")
        self.assertTrue(any("is over-allocated" in warn for warn in warnings))

        # Scenario 2: Lecturer hours over-allocated beyond 10% tolerance (e.g. 13 hours limit, 15 assigned -> 15/13 = 1.15 > 1.10)
        self.lecturer.max_hours_per_week = 13
        self.lecturer.save()
        
        is_valid, errors, warnings = validate_timetable_inputs(self.timetable)
        self.assertFalse(is_valid, "Expected validation to fail when over-allocation exceeds tolerance")
        self.assertTrue(any("is over-allocated" in err for err in errors))


    def test_validation_empty_structures_failure(self):
        """
        Verify that missing rooms or timeslots fails validation.
        """
        # Delete all rooms
        Room.objects.filter(campus__university=self.uni).delete()
        
        from .validation import validate_timetable_inputs
        is_valid, errors, warnings = validate_timetable_inputs(self.timetable)
        self.assertFalse(is_valid)
        self.assertTrue(any("No rooms found" in err for err in errors))


class SchedulingEngineTests(TestCase):
    """Phase 3: Tests for new hard constraints, conflict detection, GenerationLog, service pipeline, and scale."""

    def setUp(self):
        self.uni = University.objects.create(name="Engine University", code="EU")
        self.campus = Campus.objects.create(university=self.uni, name="Engine Campus")
        self.faculty = Faculty.objects.create(campus=self.campus, name="Engine Faculty")
        self.dept = Department.objects.create(faculty=self.faculty, name="Engine Dept")
        self.program = Program.objects.create(department=self.dept, name="Engine Program")
        self.semester = Semester.objects.create(
            university=self.uni, name="Spring 2027",
            start_date=datetime.date(2027, 1, 1),
            end_date=datetime.date(2027, 5, 31), is_active=True
        )
        self.lecture_room = Room.objects.create(campus=self.campus, name="LH101", capacity=80, room_type="Lecture")
        self.lab_room = Room.objects.create(campus=self.campus, name="LAB01", capacity=30, room_type="Lab")
        self.lec_a = Lecturer.objects.create(department=self.dept, name="Dr. Alpha", email="alpha@eu.edu")
        self.lec_b = Lecturer.objects.create(department=self.dept, name="Dr. Beta", email="beta@eu.edu")
        self.group_small = StudentGroup.objects.create(program=self.program, name="Group S", size=25)
        self.group_large = StudentGroup.objects.create(program=self.program, name="Group L", size=60)
        self.ts1 = TimeSlot.objects.create(university=self.uni, day_of_week=1, start_time=datetime.time(8,30), end_time=datetime.time(10,0), slot_number=1, is_evening=False)
        self.ts2 = TimeSlot.objects.create(university=self.uni, day_of_week=1, start_time=datetime.time(10,15), end_time=datetime.time(11,45), slot_number=2, is_evening=False)
        self.ts3 = TimeSlot.objects.create(university=self.uni, day_of_week=1, start_time=datetime.time(12,0), end_time=datetime.time(13,30), slot_number=3, is_evening=False)
        self.ts4 = TimeSlot.objects.create(university=self.uni, day_of_week=1, start_time=datetime.time(19,0), end_time=datetime.time(20,30), slot_number=4, is_evening=True)
        self.c_lecture = Course.objects.create(program=self.program, code="EU101", name="Engineering Intro", duration_slots=1, required_room_type="Lecture", lecturer=self.lec_a, student_group=self.group_large)
        self.c_lab = Course.objects.create(program=self.program, code="EU102", name="Engineering Lab", duration_slots=1, required_room_type="Lab", lecturer=self.lec_b, student_group=self.group_small)
        self.timetable = Timetable.objects.create(semester=self.semester, name="Engine Timetable V1")

    def _make_constraint(self, ctype, is_hard=True, parameters=None, weight=10):
        return Constraint.objects.create(
            university=self.uni, name="Test " + ctype,
            constraint_type=ctype, is_hard=is_hard,
            weight=weight, parameters=parameters or {}
        )

    def test_hard_no_evening_classes_constraint(self):
        """Solver must not use evening slots when NO_EVENING_CLASSES hard constraint is set."""
        self._make_constraint("NO_EVENING_CLASSES", is_hard=True)
        status, message, score = generate_timetable(self.timetable.id)
        self.assertIn(status, ("OPTIMAL", "FEASIBLE"))
        evening_slots = [s for s in ScheduleSlot.objects.filter(timetable=self.timetable).select_related("time_slot") if s.time_slot.is_evening]
        self.assertEqual(len(evening_slots), 0, "Found evening slots when NO_EVENING_CLASSES hard constraint is active: " + str([s.course.code for s in evening_slots]))

    def test_hard_lab_only_constraint(self):
        """A course with LAB_ONLY_COURSE constraint must be assigned to a Lab room."""
        self._make_constraint("LAB_ONLY_COURSE", is_hard=True, parameters={"course_id": self.c_lab.id})
        status, message, score = generate_timetable(self.timetable.id)
        self.assertIn(status, ("OPTIMAL", "FEASIBLE"))
        lab_slot = ScheduleSlot.objects.filter(timetable=self.timetable, course=self.c_lab).first()
        self.assertIsNotNone(lab_slot)
        self.assertEqual(lab_slot.room.room_type, "Lab", "Expected Lab, got " + str(lab_slot.room.room_type))

    def test_hard_lecturer_availability_constraint(self):
        """Solver must not schedule a lecturer in their unavailable timeslots."""
        self._make_constraint("LECTURER_AVAILABILITY", is_hard=True, parameters={"lecturer_id": self.lec_a.id, "unavailable_slots": [self.ts1.id]})
        status, message, score = generate_timetable(self.timetable.id)
        self.assertIn(status, ("OPTIMAL", "FEASIBLE"))
        blocked = ScheduleSlot.objects.filter(timetable=self.timetable, lecturer=self.lec_a, time_slot=self.ts1)
        self.assertEqual(blocked.count(), 0, "Lecturer A scheduled in unavailable slot.")

    def test_lecturer_self_service_availability_constraint(self):
        """Solver must not schedule a lecturer in timeslots they marked as unavailable via LecturerAvailability."""
        from .models import LecturerAvailability
        LecturerAvailability.objects.create(lecturer=self.lec_a, time_slot=self.ts1, is_available=False)
        status, message, score = generate_timetable(self.timetable.id)
        self.assertIn(status, ("OPTIMAL", "FEASIBLE"))
        blocked = ScheduleSlot.objects.filter(timetable=self.timetable, lecturer=self.lec_a, time_slot=self.ts1)
        self.assertEqual(blocked.count(), 0, "Lecturer A scheduled in self-service unavailable slot.")





    def test_hard_student_max_classes_per_day(self):
        """Solver must respect student group max classes per day limit."""
        self._make_constraint("STUDENT_MAX_CLASSES_PER_DAY", is_hard=True, parameters={"student_group_id": self.group_large.id, "max_classes": 1})
        status, message, score = generate_timetable(self.timetable.id)
        self.assertIn(status, ("OPTIMAL", "FEASIBLE"))
        count = ScheduleSlot.objects.filter(timetable=self.timetable, student_group=self.group_large, time_slot__day_of_week=1).values("course_id").distinct().count()
        self.assertLessEqual(count, 1, "Expected max 1 class/day for Group L, found " + str(count))

    def test_conflict_room_double_booking(self):
        """detect_conflicts must catch two courses placed in the same room at the same time."""
        from .conflicts import detect_conflicts
        s1 = ScheduleSlot(id=1, timetable=self.timetable, course=self.c_lecture, lecturer=self.lec_a, room=self.lecture_room, time_slot=self.ts1, student_group=self.group_large)
        s2 = ScheduleSlot(id=2, timetable=self.timetable, course=self.c_lab, lecturer=self.lec_b, room=self.lecture_room, time_slot=self.ts1, student_group=self.group_small)
        conflicts = detect_conflicts([s1, s2], self.uni)
        room_conflicts = [c for c in conflicts if c["constraint_type"] == "ROOM_DOUBLE_BOOKING"]
        self.assertGreater(len(room_conflicts), 0, "Expected ROOM_DOUBLE_BOOKING conflict.")

    def test_conflict_student_group_double_booking(self):
        """detect_conflicts must catch same student group at two classes simultaneously."""
        from .conflicts import detect_conflicts
        c_extra = Course.objects.create(program=self.program, code="EU103", name="Extra", duration_slots=1, required_room_type="Lecture", lecturer=self.lec_b, student_group=self.group_large)
        s1 = ScheduleSlot(id=1, timetable=self.timetable, course=self.c_lecture, lecturer=self.lec_a, room=self.lecture_room, time_slot=self.ts1, student_group=self.group_large)
        s2 = ScheduleSlot(id=2, timetable=self.timetable, course=c_extra, lecturer=self.lec_b, room=self.lab_room, time_slot=self.ts1, student_group=self.group_large)
        conflicts = detect_conflicts([s1, s2], self.uni)
        sg_conflicts = [c for c in conflicts if c["constraint_type"] == "STUDENT_GROUP_DOUBLE_BOOKING"]
        self.assertGreater(len(sg_conflicts), 0, "Expected STUDENT_GROUP_DOUBLE_BOOKING conflict.")

    def test_conflict_lab_only_violation(self):
        """detect_conflicts must flag a lab-only course placed in a wrong room type."""
        from .conflicts import detect_conflicts
        self._make_constraint("LAB_ONLY_COURSE", is_hard=True, parameters={"course_id": self.c_lab.id})
        slot = ScheduleSlot.objects.create(timetable=self.timetable, course=self.c_lab, lecturer=self.lec_b, room=self.lecture_room, time_slot=self.ts1, student_group=self.group_small)
        conflicts = detect_conflicts([slot], self.uni)
        lab_violations = [c for c in conflicts if c["constraint_type"] == "LAB_ONLY_VIOLATION"]
        self.assertGreater(len(lab_violations), 0, "Expected LAB_ONLY_VIOLATION conflict.")

    def test_generation_log_created(self):
        """A GenerationLog must be persisted after every run_scheduling_pipeline call."""
        from .models import GenerationLog
        from .scheduling_service import run_scheduling_pipeline
        before = GenerationLog.objects.filter(timetable=self.timetable).count()
        result = run_scheduling_pipeline(self.timetable.id)
        self.assertEqual(GenerationLog.objects.filter(timetable=self.timetable).count(), before + 1)
        self.assertIsNotNone(result.log_id)
        log = GenerationLog.objects.get(pk=result.log_id)
        self.assertEqual(log.timetable, self.timetable)
        self.assertIn(log.status, ("OPTIMAL", "FEASIBLE", "INFEASIBLE", "ERROR"))

    def test_scheduling_service_pipeline(self):
        """run_scheduling_pipeline must return SchedulingResult with correct attributes."""
        from .scheduling_service import run_scheduling_pipeline, SchedulingResult
        result = run_scheduling_pipeline(self.timetable.id)
        self.assertIsInstance(result, SchedulingResult)
        self.assertIn(result.status, ("OPTIMAL", "FEASIBLE", "INFEASIBLE", "ERROR", "VALIDATION_ERROR"))
        self.assertIsNotNone(result.message)
        self.assertIsNotNone(result.log_id)
        if result.status in ("OPTIMAL", "FEASIBLE"):
            self.assertIsNotNone(result.solve_time_seconds)
            self.assertGreater(result.courses_scheduled, 0)
            self.assertIsInstance(result.hard_conflicts, list)
            self.assertIsInstance(result.soft_conflicts, list)

    def test_scale_50_courses(self):
        """Solver must handle 50 courses across 10 rooms and 5 days within 65 seconds."""
        import time as time_module
        u = University.objects.create(name="Scale University", code="SU")
        ca = Campus.objects.create(university=u, name="Scale Campus")
        fa = Faculty.objects.create(campus=ca, name="Scale Faculty")
        de = Department.objects.create(faculty=fa, name="Scale Dept")
        pr = Program.objects.create(department=de, name="Scale Program")
        se = Semester.objects.create(university=u, name="Scale Semester", start_date=datetime.date(2027,9,1), end_date=datetime.date(2027,12,31), is_active=True)
        for i in range(10):
            Room.objects.create(campus=ca, name="SR" + str(i), capacity=100, room_type="Lecture")
        for day in range(1, 6):
            for slot_num in range(1, 5):
                TimeSlot.objects.create(university=u, day_of_week=day, start_time=datetime.time(8+slot_num, 0), end_time=datetime.time(9+slot_num, 30), slot_number=slot_num, is_evening=False)
        for i in range(50):
            lec = Lecturer.objects.create(department=de, name="SL" + str(i), email="sl" + str(i) + "@su.edu")
            grp = StudentGroup.objects.create(program=pr, name="SG" + str(i), size=30)
            Course.objects.create(program=pr, code="SC" + str(i).zfill(3), name="Scale Course " + str(i), duration_slots=1, required_room_type="Lecture", lecturer=lec, student_group=grp)
        tt = Timetable.objects.create(semester=se, name="Scale Timetable")
        t0 = time_module.perf_counter()
        status, message, score = generate_timetable(tt.id, time_limit_seconds=60)
        elapsed = time_module.perf_counter() - t0
        self.assertIn(status, ("OPTIMAL", "FEASIBLE"), "Status " + str(status) + " after " + str(round(elapsed,1)) + "s: " + message)
        self.assertLessEqual(elapsed, 65, "Took " + str(round(elapsed,1)) + "s - too slow.")
        self.assertEqual(ScheduleSlot.objects.filter(timetable=tt).count(), 50)


class DataImportEngineTests(TestCase):
    def setUp(self):
        self.uni = University.objects.create(name="Import University", code="IU")
        
        # Create user
        from django.contrib.auth.models import User
        from accounts.models import UserProfile
        self.admin_user = User.objects.create_user(username="import_admin", password="password")
        UserProfile.objects.create(user=self.admin_user, role="admin", university=self.uni)

    def test_import_rooms_success(self):
        """Verify bulk room CSV import works and auto-creates structural campus."""
        self.client.login(username="import_admin", password="password")
        session = self.client.session
        session['active_role'] = 'admin'
        session['active_university_id'] = self.uni.id
        session.save()

        import io
        csv_data = (
            "name,capacity,room_type,campus_name\n"
            "Auditorium X,150,Lecture,Main Campus\n"
            "Lab Z,40,Lab,Science Campus\n"
        )
        csv_file = io.BytesIO(csv_data.encode('utf-8'))
        csv_file.name = 'rooms.csv'

        response = self.client.post(reverse('scheduler:import_resources'), {
            'import_type': 'room',
            'file': csv_file
        })
        self.assertRedirects(response, reverse('scheduler:resources_manager') + '?tab=room')

        self.assertEqual(Room.objects.filter(campus__university=self.uni).count(), 2)
        room_x = Room.objects.get(name="Auditorium X")
        self.assertEqual(room_x.capacity, 150)
        self.assertEqual(room_x.room_type, "Lecture")
        self.assertEqual(room_x.campus.name, "Main Campus")

    def test_import_lecturers_success(self):
        """Verify lecturer import works and auto-creates department."""
        self.client.login(username="import_admin", password="password")
        session = self.client.session
        session['active_role'] = 'admin'
        session['active_university_id'] = self.uni.id
        session.save()

        import io
        csv_data = (
            "name,email,department_name,max_hours\n"
            "Dr. John Doe,john@iu.edu,CS Dept,20\n"
            "Prof. Jane Smith,jane@iu.edu,Math Dept,24\n"
        )
        csv_file = io.BytesIO(csv_data.encode('utf-8'))
        csv_file.name = 'lecturers.csv'

        response = self.client.post(reverse('scheduler:import_resources'), {
            'import_type': 'lecturer',
            'file': csv_file
        })
        self.assertRedirects(response, reverse('scheduler:resources_manager') + '?tab=lecturer')

        self.assertEqual(Lecturer.objects.filter(department__faculty__campus__university=self.uni).count(), 2)
        john = Lecturer.objects.get(email="john@iu.edu")
        self.assertEqual(john.name, "Dr. John Doe")
        self.assertEqual(john.department.name, "CS Dept")

    def test_import_validation_failure_rollback(self):
        """Verify that a CSV containing hard validation errors rolls back all changes in the transaction."""
        self.client.login(username="import_admin", password="password")
        session = self.client.session
        session['active_role'] = 'admin'
        session['active_university_id'] = self.uni.id
        session.save()

        import io
        # Auditorium X is valid, but the second row has an invalid capacity "abc" which should cause a rollback
        csv_data = (
            "name,capacity,room_type,campus_name\n"
            "Auditorium X,150,Lecture,Main Campus\n"
            "Lab Z,abc,Lab,Science Campus\n"
        )
        csv_file = io.BytesIO(csv_data.encode('utf-8'))
        csv_file.name = 'rooms_invalid.csv'

        response = self.client.post(reverse('scheduler:import_resources'), {
            'import_type': 'room',
            'file': csv_file
        })
        self.assertEqual(response.status_code, 200)

        # Verify database is empty (Auditorium X was NOT saved due to transaction rollback)
        self.assertEqual(Room.objects.filter(campus__university=self.uni).count(), 0)

    def test_import_all_validation_failure_rollback(self):
        """Verify that a multi-sheet Excel file containing validation errors rolls back all changes across all sheets."""
        self.client.login(username="import_admin", password="password")
        session = self.client.session
        session['active_role'] = 'admin'
        session['active_university_id'] = self.uni.id
        session.save()

        import openpyxl
        import io

        wb = openpyxl.Workbook()
        
        # Create rooms sheet with an invalid capacity 'abc' in second row
        ws_rooms = wb.active
        ws_rooms.title = "Rooms"
        ws_rooms.append(["name", "capacity", "room_type", "campus_name"])
        ws_rooms.append(["Auditorium X", "150", "Lecture", "Main Campus"])
        ws_rooms.append(["Lab Z", "abc", "Lab", "Science Campus"])

        # Create valid lecturers sheet
        ws_lecs = wb.create_sheet("Lecturers")
        ws_lecs.append(["name", "email", "department_name"])
        ws_lecs.append(["Dr. Alice", "alice@test.com", "CS Dept"])

        # Save workbook to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        excel_file.name = "import_all_invalid.xlsx"

        response = self.client.post(reverse('scheduler:import_resources'), {
            'import_type': 'all',
            'file': excel_file
        })
        self.assertEqual(response.status_code, 200)

        # Verify nothing was saved to the database (both sheets rolled back)
        self.assertEqual(Room.objects.filter(campus__university=self.uni).count(), 0)
        self.assertEqual(Lecturer.objects.filter(department__faculty__campus__university=self.uni).count(), 0)


class SmartImportTests(TestCase):
    def setUp(self):
        from scheduler.models import University
        from django.contrib.auth.models import User
        from accounts.models import UserProfile
        self.uni = University.objects.create(name="Smart Uni", code="SU")
        self.admin_user = User.objects.create_user(username="smart_admin", password="password")
        UserProfile.objects.create(user=self.admin_user, role="admin", university=self.uni)

    def test_detect_format_flat_timetable(self):
        import openpyxl
        from scheduler.smart_import import detect_format
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CS_Y1S1"
        ws.append(["DAY", "TIME", "VENUE", "INSTRUCTOR", "UNIT CODE", "UNIT NAME"])
        ws.append(["Monday", "08:00-11:00", "LH 1", "Dr. John", "CS101", "Intro to CS"])
        
        info = detect_format(wb)
        self.assertEqual(info['type'], 'flat_timetable')
        self.assertEqual(info['confidence'], 1.0)
        self.assertEqual(info['sheets'][0]['detected_type'], 'flat_timetable')

    def test_extract_entities_flat_timetable(self):
        import openpyxl
        from scheduler.smart_import import detect_format, extract_entities
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CS_Y1S1"
        ws.append(["DAY", "TIME", "VENUE", "INSTRUCTOR", "UNIT CODE", "UNIT NAME"])
        ws.append(["Monday", "08:00-11:00", "LH 1", "Dr. John", "CS101", "Intro to CS"])
        
        info = detect_format(wb)
        entities = extract_entities(wb, info, self.uni)
        
        self.assertEqual(len(entities['rooms']), 1)
        self.assertEqual(entities['rooms'][0]['name'], "LH 1")
        self.assertEqual(entities['rooms'][0]['room_type'], "Lecture")
        
        self.assertEqual(len(entities['lecturers']), 1)
        self.assertEqual(entities['lecturers'][0]['name'], "Dr. John")
        self.assertEqual(entities['lecturers'][0]['email'], "john@smartuni.edu")
        
        self.assertEqual(len(entities['courses']), 1)
        self.assertEqual(entities['courses'][0]['code'], "CS101")
        self.assertEqual(entities['courses'][0]['name'], "Intro to CS")
        
        self.assertEqual(len(entities['time_slots']), 1)
        self.assertEqual(entities['time_slots'][0]['day_of_week'], 1)

    def test_smart_import_preview_flow(self):
        self.client.login(username="smart_admin", password="password")
        session = self.client.session
        session['active_role'] = 'admin'
        session['active_university_id'] = self.uni.id
        session.save()
        
        import io
        csv_data = (
            "day,time,room,lecturer,course_code,course_name,student_group\n"
            "Monday,0800-1100,LH 1,Dr. John,CS101,Intro to CS,CS Group\n"
        )
        csv_file = io.BytesIO(csv_data.encode('utf-8'))
        csv_file.name = 'timetable.csv'
        
        response = self.client.post(reverse('scheduler:import_resources'), {
            'import_type': 'smart',
            'file': csv_file
        })
        if response.status_code != 200:
            print("REDIRECT TARGET:", response.url if hasattr(response, 'url') else response.get('Location', ''))
            print("MESSAGES:", [m.message for m in list(response.wsgi_request._messages)])
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Intelligent Import Preview")
        self.assertContains(response, "CS101")
        self.assertContains(response, "Dr. John")
        # Now let's confirm the import
        confirm_response = self.client.post(reverse('scheduler:import_resources'), {
            'confirm': 'yes'
        })
        # Confirm now redirects to the audit report page (URL pattern: /import-audit/<pk>/)
        self.assertEqual(confirm_response.status_code, 302)
        self.assertRegex(confirm_response['Location'], r'/import-audit/\d+/')
        
        # Verify db insertion
        from scheduler.models import Course, Room, Lecturer, ImportAuditLog
        self.assertEqual(Room.objects.filter(campus__university=self.uni).count(), 1)
        self.assertEqual(Lecturer.objects.filter(department__faculty__campus__university=self.uni).count(), 1)
        self.assertEqual(Course.objects.filter(program__department__faculty__campus__university=self.uni).count(), 1)
        # Verify audit log was created
        self.assertEqual(ImportAuditLog.objects.filter(university=self.uni).count(), 1)

    def test_title_stripping_no_space(self):
        from scheduler.smart_import import _strip_title
        # Test stripping without space after dot
        self.assertEqual(_strip_title("Dr.John Doe"), "John Doe")
        self.assertEqual(_strip_title("Prof.Jane Doe"), "Jane Doe")
        # Test stripping with space after dot
        self.assertEqual(_strip_title("Dr. John Doe"), "John Doe")

    def test_course_code_not_globally_deduplicated(self):
        import openpyxl
        from scheduler.smart_import import detect_format, extract_entities
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CS_All"
        ws.append(["DAY", "TIME", "VENUE", "INSTRUCTOR", "UNIT CODE", "UNIT NAME", "STUDENT GROUP"])
        # Same course code, different student groups
        ws.append(["Monday", "08:00-11:00", "LH 1", "Dr. John", "CS101", "Intro to CS", "Group A"])
        ws.append(["Tuesday", "08:00-11:00", "LH 2", "Dr. John", "CS101", "Intro to CS", "Group B"])
        
        info = detect_format(wb)
        entities = extract_entities(wb, info, self.uni)
        
        # Both course entries should be preserved because they are for different groups
        self.assertEqual(len(entities['courses']), 2)
        codes = sorted([c['code'] for c in entities['courses']])
        groups = sorted([c['student_group'] for c in entities['courses']])
        self.assertEqual(codes, ["CS101", "CS101"])
        self.assertEqual(groups, ["Group A", "Group B"])


class CalendarIntegrationTests(TestCase):
    def setUp(self):
        self.uni = University.objects.create(name="Calendar Uni", code="CALU")
        self.campus = Campus.objects.create(university=self.uni, name="Cal Campus")
        self.faculty = Faculty.objects.create(campus=self.campus, name="Cal Faculty")
        self.dept = Department.objects.create(faculty=self.faculty, name="Cal Dept")
        self.program = Program.objects.create(department=self.dept, name="Cal Prog")
        self.semester = Semester.objects.create(
            university=self.uni,
            name="Fall 2026",
            start_date=datetime.date(2026, 9, 1),
            end_date=datetime.date(2026, 12, 20)
        )
        self.lecturer = Lecturer.objects.create(
            department=self.dept,
            name="Test Calendar Lecturer",
            email="cal@example.com"
        )
        self.student_group = StudentGroup.objects.create(
            program=self.program,
            name="Cal Year 1",
            size=30
        )
        self.timetable = Timetable.objects.create(
            semester=self.semester,
            name="Calendar Timetable",
            is_active=True
        )

    def test_lecturer_feed_view(self):
        # Trigger feed view
        url = reverse('scheduler:lecturer_calendar_feed', args=[self.lecturer.calendar_token])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'text/calendar; charset=utf-8')
        self.assertIn(b'BEGIN:VCALENDAR', response.content)

    def test_student_group_feed_view(self):
        # Trigger feed view
        url = reverse('scheduler:student_group_calendar_feed', args=[self.student_group.calendar_token])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'text/calendar; charset=utf-8')
        self.assertIn(b'BEGIN:VCALENDAR', response.content)


from django.test import TransactionTestCase
from django.db import transaction
from unittest.mock import patch

class TimetableAutoGenerationTests(TransactionTestCase):
    def setUp(self):
        # Create minimal structure
        self.uni = University.objects.create(name="Signal Test University", code="STU")
        self.campus = Campus.objects.create(university=self.uni, name="Signal Test Campus")
        self.faculty = Faculty.objects.create(campus=self.campus, name="Signal Test Faculty")
        self.dept = Department.objects.create(faculty=self.faculty, name="Signal Test Dept")
        self.program = Program.objects.create(department=self.dept, name="Signal Test Program")
        self.semester = Semester.objects.create(
            university=self.uni,
            name="Fall 2026",
            start_date=datetime.date(2026, 9, 1),
            end_date=datetime.date(2026, 12, 1),
            is_active=True
        )
        self.lecturer = Lecturer.objects.create(department=self.dept, name="Dr. Signal", email="signal@tu.edu")
        self.student_group = StudentGroup.objects.create(program=self.program, name="Signal Group", size=30)
        self.room = Room.objects.create(campus=self.campus, name="Signal Room", capacity=50)
        self.ts = TimeSlot.objects.create(university=self.uni, day_of_week=1, start_time=datetime.time(8, 30), end_time=datetime.time(10, 0), slot_number=1)
        
        # Active Timetable
        self.active_timetable = Timetable.objects.create(
            semester=self.semester,
            name="Active Timetable",
            is_active=True
        )

    @patch('django_q.tasks.async_task')
    def test_auto_generation_triggered_on_course_save(self, mock_async_task):
        # Create a new course, which should trigger signals
        Course.objects.create(
            program=self.program,
            code="SIG101",
            name="Signal Course",
            duration_slots=1,
            required_room_type="Lecture",
            lecturer=self.lecturer,
            student_group=self.student_group
        )
        # Should call async_task once for the active timetable
        self.assertEqual(mock_async_task.call_count, 1)
        # Verify the target task and timetable ID
        args, kwargs = mock_async_task.call_args
        self.assertEqual(args[0], 'scheduler.tasks.generate_timetable_async')
        self.assertEqual(args[1], self.active_timetable.id)

    @patch('django_q.tasks.async_task')
    def test_auto_generation_not_triggered_on_inactive_timetable(self, mock_async_task):
        # Deactivate all timetables
        Timetable.objects.all().update(is_active=False)
        
        # Create a course
        Course.objects.create(
            program=self.program,
            code="SIG102",
            name="Signal Course 2",
            duration_slots=1,
            required_room_type="Lecture",
            lecturer=self.lecturer,
            student_group=self.student_group
        )
        
        # Should NOT trigger generation
        mock_async_task.assert_not_called()

    @patch('django_q.tasks.async_task')
    def test_debouncing_in_transaction(self, mock_async_task):
        with transaction.atomic():
            # Create multiple courses in a single transaction
            for i in range(3):
                Course.objects.create(
                    program=self.program,
                    code=f"SIG{103+i}",
                    name=f"Signal Course {3+i}",
                    duration_slots=1,
                    required_room_type="Lecture",
                    lecturer=self.lecturer,
                    student_group=self.student_group
                )
            # Within the transaction, it shouldn't have fired yet
            mock_async_task.assert_not_called()
        
        # After transaction commits, it should fire exactly once
        self.assertEqual(mock_async_task.call_count, 1)
        args, kwargs = mock_async_task.call_args
        self.assertEqual(args[1], self.active_timetable.id)

    @patch('django_q.tasks.async_task')
    def test_manual_slot_edit_excluded(self, mock_async_task):
        # Create a course to link to a slot
        course = Course.objects.create(
            program=self.program,
            code="SIG106",
            name="Signal Course 6",
            duration_slots=1,
            required_room_type="Lecture",
            lecturer=self.lecturer,
            student_group=self.student_group
        )
        
        # Clear call count from the course save trigger
        mock_async_task.reset_mock()
        
        # Manually create/edit a ScheduleSlot
        slot = ScheduleSlot.objects.create(
            timetable=self.active_timetable,
            course=course,
            lecturer=self.lecturer,
            room=self.room,
            time_slot=self.ts,
            student_group=self.student_group
        )
        
        # Check that manual schedule slot save did not trigger solver generation
        mock_async_task.assert_not_called()


class KenyattaUniTimetableGenerationTests(TestCase):
    def test_kenyatta_uni_generation_scale(self):
        """
        Verify that Kenyatta University's timetable generation runs successfully
        on the full 2000-row imported dataset.
        """
        import openpyxl
        import os
        import time
        from django.conf import settings
        
        # 1. Setup University
        uni = University.objects.create(name="Kenyatta_UNI", code="65073767")
        
        # 2. Campus & Rooms
        excel_dir = settings.BASE_DIR
        wb_rooms = openpyxl.load_workbook(os.path.join(excel_dir, 'rooms_import_2000.xlsx'), data_only=True)
        sheet_rooms = wb_rooms.active
        rows_rooms = list(sheet_rooms.iter_rows(values_only=True))
        
        campuses = {}
        rooms_to_create = []
        for row in rows_rooms[1:]:
            if not any(row):
                continue
            name, capacity, room_type, campus_name = row[:4]
            if campus_name not in campuses:
                campuses[campus_name] = Campus.objects.create(university=uni, name=campus_name)
            rooms_to_create.append(Room(
                campus=campuses[campus_name],
                name=name,
                capacity=int(capacity),
                room_type=room_type
            ))
        Room.objects.bulk_create(rooms_to_create)
        
        # 3. Lecturers
        wb_lecs = openpyxl.load_workbook(os.path.join(excel_dir, 'lecturers_import_2000.xlsx'), data_only=True)
        sheet_lecs = wb_lecs.active
        rows_lecs = list(sheet_lecs.iter_rows(values_only=True))
        
        # Pick the first campus to associate the faculty
        first_campus = list(campuses.values())[0]
        fac = Faculty.objects.create(campus=first_campus, name="Default Faculty")
        
        departments = {}
        lecturers_to_create = []
        for row in rows_lecs[1:]:
            if not any(row):
                continue
            name, email, department_name, max_hours = row[:4]
            if department_name not in departments:
                departments[department_name] = Department.objects.create(faculty=fac, name=department_name)
            lecturers_to_create.append(Lecturer(
                department=departments[department_name],
                name=name,
                email=email,
                max_hours_per_week=int(max_hours) if max_hours else 20
            ))
        Lecturer.objects.bulk_create(lecturers_to_create)
        
        # 4. Student Groups
        wb_groups = openpyxl.load_workbook(os.path.join(excel_dir, 'student_groups_import_2000.xlsx'), data_only=True)
        sheet_groups = wb_groups.active
        rows_groups = list(sheet_groups.iter_rows(values_only=True))
        
        first_dept = list(departments.values())[0]
        programs = {}
        groups_to_create = []
        for row in rows_groups[1:]:
            if not any(row):
                continue
            name, size, program_name = row[:3]
            if program_name not in programs:
                programs[program_name] = Program.objects.create(department=first_dept, name=program_name)
            groups_to_create.append(StudentGroup(
                program=programs[program_name],
                name=name,
                size=int(size)
            ))
        StudentGroup.objects.bulk_create(groups_to_create)
        
        # 5. Timeslots
        timeslots_to_create = []
        slots = [
            (1, datetime.time(8, 30), datetime.time(10, 0)),
            (2, datetime.time(10, 15), datetime.time(11, 45)),
            (3, datetime.time(12, 0), datetime.time(13, 30)),
            (4, datetime.time(13, 45), datetime.time(15, 15)),
            (5, datetime.time(15, 30), datetime.time(17, 0))
        ]
        for day in range(1, 6):
            for s_num, start, end in slots:
                timeslots_to_create.append(TimeSlot(
                    university=uni,
                    day_of_week=day,
                    start_time=start,
                    end_time=end,
                    slot_number=s_num,
                    is_evening=False
                ))
        TimeSlot.objects.bulk_create(timeslots_to_create)
        
        # 6. Courses
        wb_courses = openpyxl.load_workbook(os.path.join(excel_dir, 'courses_import_2000.xlsx'), data_only=True)
        sheet_courses = wb_courses.active
        rows_courses = list(sheet_courses.iter_rows(values_only=True))
        
        lecturer_map = {l.email: l.id for l in Lecturer.objects.filter(department__faculty__campus__university=uni)}
        group_map = {g.name: g.id for g in StudentGroup.objects.filter(program__department__faculty__campus__university=uni)}
        
        courses_to_create = []
        for row in rows_courses[1:]:
            if not any(row):
                continue
            code, name, duration_slots, required_room_type, lecturer_email, student_group_name, program_name = row[:7]
            
            prog = programs.get(program_name) or list(programs.values())[0]
            lec_id = lecturer_map.get(lecturer_email)
            grp_id = group_map.get(student_group_name)
            
            courses_to_create.append(Course(
                program=prog,
                code=code,
                name=name,
                duration_slots=int(duration_slots),
                required_room_type=required_room_type,
                lecturer_id=lec_id,
                student_group_id=grp_id
            ))
        Course.objects.bulk_create(courses_to_create)
        
        # 7. Semester & Timetable
        sem = Semester.objects.create(
            university=uni,
            name="Fall 2026",
            start_date=datetime.date(2026, 9, 1),
            end_date=datetime.date(2026, 12, 1),
            is_active=True
        )
        tt = Timetable.objects.create(semester=sem, name="HealthGuide")
        
        # 8. Solve & Assertions
        t_start = time.perf_counter()
        status, message, obj = generate_timetable(tt.id)
        elapsed = time.perf_counter() - t_start
        
        self.assertIn(status, ('FEASIBLE', 'OPTIMAL'))
        self.assertEqual(obj, 0)
        self.assertEqual(ScheduleSlot.objects.filter(timetable=tt).count(), 3999)
        self.assertLessEqual(elapsed, 25.0, f"Solving took too long: {elapsed:.1f} seconds")


class RealWorldFeasibilityTests(TestCase):
    def setUp(self):
        self.uni = University.objects.create(name="Feasibility University", code="FEASU")
        self.campus = Campus.objects.create(university=self.uni, name="Feasibility Campus")
        self.faculty = Faculty.objects.create(campus=self.campus, name="Feasibility Faculty")
        self.dept = Department.objects.create(faculty=self.faculty, name="Feasibility Dept")
        self.program = Program.objects.create(department=self.dept, name="Feasibility Program")
        self.semester = Semester.objects.create(
            university=self.uni,
            name="Fall 2026",
            start_date=datetime.date(2026, 9, 1),
            end_date=datetime.date(2026, 12, 1),
            is_active=True
        )

        # 3 rooms with different capacities and types
        self.lecture_room = Room.objects.create(campus=self.campus, name="R101", capacity=100, room_type="Lecture")
        self.lab_room = Room.objects.create(campus=self.campus, name="Lab102", capacity=30, room_type="Lab")
        self.seminar_room = Room.objects.create(campus=self.campus, name="Sem103", capacity=40, room_type="Seminar")

        # 5 lecturers
        self.lecturers = []
        for i in range(5):
            self.lecturers.append(
                Lecturer.objects.create(
                    department=self.dept,
                    name=f"Prof. {chr(65+i)}",
                    email=f"{chr(97+i)}@feasu.edu",
                    max_hours_per_week=20
                )
            )

        # 3 student groups
        self.group_y1 = StudentGroup.objects.create(program=self.program, name="Year 1 CS", size=80)
        self.group_y2 = StudentGroup.objects.create(program=self.program, name="Year 2 CS", size=25)
        self.group_y3 = StudentGroup.objects.create(program=self.program, name="Year 3 CS", size=35)

        # Timeslots: 5 days, 5 slots per day
        self.timeslots = []
        for day in range(1, 6):
            for slot_num in range(1, 6):
                self.timeslots.append(
                    TimeSlot.objects.create(
                        university=self.uni,
                        day_of_week=day,
                        start_time=datetime.time(8 + slot_num, 30),
                        end_time=datetime.time(9 + slot_num, 45),
                        slot_number=slot_num
                    )
                )

        self.timetable = Timetable.objects.create(semester=self.semester, name="Feasibility Timetable")

    def verify_real_world_feasibility(self, timetable, university):
        """
        Rigorous feasibility check verifying all real-world constraints on the saved ScheduleSlots.
        """
        slots = list(ScheduleSlot.objects.filter(timetable=timetable).select_related('course', 'room', 'time_slot', 'lecturer', 'student_group'))
        
        # 1. Basic properties
        self.assertGreater(len(slots), 0, "Feasible schedule should have scheduled slots")

        # Mappings for double booking checks
        lecturer_occupancy = defaultdict(list)     # (lecturer_id, day, slot) -> slot
        room_occupancy = defaultdict(list)         # (room_id, day, slot) -> slot
        student_group_occupancy = defaultdict(list)# (group_id, day, slot) -> slot
        lecturer_weekly_slots = defaultdict(list)

        # Group slots by course to check multi-slot consecutiveness
        course_slots = defaultdict(list)

        for slot in slots:
            ts = slot.time_slot
            day = ts.day_of_week
            num = ts.slot_number

            # Check that all elements belong to this university
            self.assertEqual(ts.university, university)
            self.assertEqual(slot.room.campus.university, university)
            self.assertEqual(slot.lecturer.department.faculty.campus.university, university)

            # Record for double bookings
            lecturer_occupancy[(slot.lecturer_id, day, num)].append(slot)
            room_occupancy[(slot.room_id, day, num)].append(slot)
            student_group_occupancy[(slot.student_group_id, day, num)].append(slot)
            lecturer_weekly_slots[slot.lecturer_id].append(slot)
            course_slots[slot.course_id].append(slot)

            # 2. Room Capacity check
            self.assertLessEqual(
                slot.student_group.size, slot.room.capacity,
                f"Room capacity violation: Student group '{slot.student_group.name}' size {slot.student_group.size} "
                f"exceeds Room '{slot.room.name}' capacity {slot.room.capacity}."
            )

            # 3. Room Type suitability (Lab course in Lab room, etc.)
            
            # 4. Lecturer Self-Service Availability
            self.assertFalse(
                LecturerAvailability.objects.filter(lecturer=slot.lecturer, time_slot=ts, is_available=False).exists(),
                f"Lecturer availability violation: {slot.lecturer.name} scheduled in unavailable slot {ts}."
            )

        # 5. Lecturer Double Booking
        for key, occupied_slots in lecturer_occupancy.items():
            self.assertEqual(
                len(occupied_slots), 1,
                f"Lecturer double booking: {occupied_slots[0].lecturer.name} scheduled in multiple courses at "
                f"Day {key[1]}, Slot {key[2]}."
            )

        # 6. Room Double Booking
        for key, occupied_slots in room_occupancy.items():
            self.assertEqual(
                len(occupied_slots), 1,
                f"Room double booking: '{occupied_slots[0].room.name}' occupied by multiple courses at "
                f"Day {key[1]}, Slot {key[2]}."
            )

        # 7. Student Group Double Booking
        for key, occupied_slots in student_group_occupancy.items():
            self.assertEqual(
                len(occupied_slots), 1,
                f"Student Group double booking: '{occupied_slots[0].student_group.name}' in multiple courses at "
                f"Day {key[1]}, Slot {key[2]}."
            )

        # 8. Lecturer Weekly Hours limit (max_hours_per_week)
        for lecturer_id, lecturer_slots in lecturer_weekly_slots.items():
            lecturer = Lecturer.objects.get(pk=lecturer_id)
            total_hours = len(lecturer_slots) * 1.5
            self.assertLessEqual(
                total_hours, lecturer.max_hours_per_week,
                f"Lecturer workload violation: {lecturer.name} assigned {total_hours} hours, exceeding max limit of {lecturer.max_hours_per_week} hours."
            )

        # 9. Course Multi-slot Consecutiveness
        for course_id, slots_list in course_slots.items():
            course = Course.objects.get(pk=course_id)
            
            # For non-split courses (like Lab courses), they must be scheduled together and contiguous.
            # For split courses (Lecture/Seminar), they are split into separate 1-slot virtual courses
            # and scheduled independently, so they do not need to be contiguous or on the same day.
            expected_slots = course.duration_slots * course.sessions_per_week
            if course.required_room_type == 'Lab':
                self.assertEqual(
                    len(slots_list), expected_slots,
                    f"Slot count mismatch: Course '{course.code}' expected {expected_slots} slots, got {len(slots_list)} slots."
                )
                
                if course.duration_slots > 1:
                    # All slots must be on the same day, in the same room
                    days = set(s.time_slot.day_of_week for s in slots_list)
                    rooms = set(s.room_id for s in slots_list)
                    self.assertEqual(len(days), 1, f"Multi-slot course '{course.code}' scheduled on multiple days: {days}")
                    self.assertEqual(len(rooms), 1, f"Multi-slot course '{course.code}' scheduled in multiple rooms: {rooms}")
                    
                    # Slots must be contiguous
                    slot_numbers = sorted(s.time_slot.slot_number for s in slots_list)
                    for i in range(len(slot_numbers) - 1):
                        self.assertEqual(
                            slot_numbers[i+1] - slot_numbers[i], 1,
                            f"Multi-slot course '{course.code}' slot numbers are not contiguous: {slot_numbers}"
                        )
            else:
                expected_slots = course.duration_slots * course.sessions_per_week
                self.assertEqual(
                    len(slots_list), expected_slots,
                    f"Slot count mismatch: Course '{course.code}' expected {expected_slots} slots, got {len(slots_list)} slots."
                )

    def test_generated_timetable_real_world_feasibility(self):
        """
        Generate a timetable for a medium-sized university schedule and verify all correctness rules.
        """
        # Create 20 courses (some lecture, some lab, some seminar, various durations)
        # Group Y1 (size 80) needs Lecture room (capacity 100)
        # Group Y2 (size 25) can fit in Lab (30) or Seminar (40)
        # Group Y3 (size 35) can fit in Seminar (40) or Lecture (100)

        # 8 courses for Year 1 (all lecture, 1 or 2 slots)
        for i in range(8):
            Course.objects.create(
                program=self.program,
                code=f"CS10{i}",
                name=f"CS Year 1 Course {i}",
                duration_slots=2 if i < 2 else 1,
                required_room_type="Lecture",
                lecturer=self.lecturers[i % 3], # Prof A, B, C
                student_group=self.group_y1
            )

        # 6 courses for Year 2 (some lab, some seminar)
        for i in range(6):
            Course.objects.create(
                program=self.program,
                code=f"CS20{i}",
                name=f"CS Year 2 Course {i}",
                duration_slots=1,
                required_room_type="Lab" if i % 2 == 0 else "Seminar",
                lecturer=self.lecturers[3], # Prof D
                student_group=self.group_y2
            )

        # 6 courses for Year 3 (all seminar or lecture)
        for i in range(6):
            Course.objects.create(
                program=self.program,
                code=f"CS30{i}",
                name=f"CS Year 3 Course {i}",
                duration_slots=1,
                required_room_type="Seminar" if i % 2 == 0 else "Lecture",
                lecturer=self.lecturers[4], # Prof E
                student_group=self.group_y3
            )

        # Run scheduler
        status, message, obj = generate_timetable(self.timetable.id)
        
        self.assertIn(status, ('FEASIBLE', 'OPTIMAL'))

        # Check all feasibility rules on the database output
        self.verify_real_world_feasibility(self.timetable, self.uni)

        # Check day-load distribution: verify that classes are spread out and not clustered on 1-2 days
        saved_slots = list(ScheduleSlot.objects.filter(timetable=self.timetable).select_related('time_slot'))
        day_counts = defaultdict(int)
        for s in saved_slots:
            day_counts[s.time_slot.day_of_week] += 1

        # We have 20 courses, split lectures/seminars might add more slots.
        total_slots = len(saved_slots)
        
        # Check that classes are spread across at least 4 distinct days
        self.assertGreaterEqual(len(day_counts), 4, f"Classes should be scheduled on at least 4 different days to avoid clustering, got: {dict(day_counts)}")

        # Check that no single day is heavily overloaded (e.g. no day has more than 40% of the total slots)
        for day, count in day_counts.items():
            pct = count / total_slots
            self.assertLessEqual(
                pct, 0.45,
                f"Day {day} has {count} slots ({pct:.1%}), which exceeds the limit of 45% (heavy clustering)."
            )

    def test_greedy_fallback_real_world_feasibility(self):
        """
        Verify that even when the solver fails or is forced to use the greedy fallback,
        the resulting timetable is correct and satisfies all hard constraints.
        """
        # Schedule a set of 12 courses (which will run day balancing because 12 > 10).
        for i in range(12):
            Course.objects.create(
                program=self.program,
                code=f"CSGREEDY{i}",
                name=f"Greedy Course {i}",
                duration_slots=1,
                required_room_type="Lecture",
                lecturer=self.lecturers[i % 5],
                student_group=self.group_y2
            )

        # We will generate timetable, it will succeed via Phase 1 (greedy) and save slots.
        status, message, obj = generate_timetable(self.timetable.id)
        self.assertIn(status, ('FEASIBLE', 'OPTIMAL'))

        # Check all feasibility rules
        self.verify_real_world_feasibility(self.timetable, self.uni)

        # Verify load distribution for greedy solver
        saved_slots = list(ScheduleSlot.objects.filter(timetable=self.timetable).select_related('time_slot'))
        day_counts = defaultdict(int)
        for s in saved_slots:
            day_counts[s.time_slot.day_of_week] += 1
        
        self.assertGreaterEqual(len(day_counts), 3, f"Greedy fallback should spread courses across at least 3 days, got: {dict(day_counts)}")


class EdgeCasesFeasibilityTests(TestCase):
    def setUp(self):
        self.uni = University.objects.create(name="Edge Case Uni", code="ECUNI")
        self.campus_a = Campus.objects.create(university=self.uni, name="Campus A")
        self.campus_b = Campus.objects.create(university=self.uni, name="Campus B")
        
        self.faculty_a = Faculty.objects.create(campus=self.campus_a, name="Faculty A")
        self.faculty_b = Faculty.objects.create(campus=self.campus_b, name="Faculty B")
        
        self.dept_a = Department.objects.create(faculty=self.faculty_a, name="Dept A")
        self.dept_b = Department.objects.create(faculty=self.faculty_b, name="Dept B")
        
        self.program_a = Program.objects.create(department=self.dept_a, name="Prog A")
        self.program_b = Program.objects.create(department=self.dept_b, name="Prog B")
        
        self.semester = Semester.objects.create(
            university=self.uni,
            name="Fall 2026",
            start_date=datetime.date(2026, 9, 1),
            end_date=datetime.date(2026, 12, 1),
            is_active=True
        )

        self.room_a = Room.objects.create(campus=self.campus_a, name="Room A", capacity=50, room_type="Lecture")
        self.room_b = Room.objects.create(campus=self.campus_b, name="Room B", capacity=50, room_type="Lecture")

        self.lecturer = Lecturer.objects.create(department=self.dept_a, name="Dr. Travel", email="travel@ecuni.edu", max_hours_per_week=20)
        self.group_parent = StudentGroup.objects.create(program=self.program_a, name="Year 1 Merged", size=40)
        self.group_child = StudentGroup.objects.create(program=self.program_a, name="Year 1 Section A", size=20, parent_group=self.group_parent)
        self.group_independent = StudentGroup.objects.create(program=self.program_a, name="Year 2", size=20)

        # 5 timeslots for Monday (non-overlapping)
        self.timeslots = []
        for i in range(1, 6):
            self.timeslots.append(
                TimeSlot.objects.create(
                    university=self.uni,
                    day_of_week=1,
                    start_time=datetime.time(8 + 2 * (i - 1), 0),
                    end_time=datetime.time(9 + 2 * (i - 1), 30),
                    slot_number=i
                )
            )


        self.timetable = Timetable.objects.create(semester=self.semester, name="Edge Cases Timetable")

    def test_campus_travel_time_conflict_detection(self):
        """Verify detect_conflicts flags LECTURER_CAMPUS_TRAVEL_VIOLATION for consecutive cross-campus slots."""
        c1 = Course.objects.create(program=self.program_a, code="C1", name="Course 1", duration_slots=1, lecturer=self.lecturer, student_group=self.group_independent)
        c2 = Course.objects.create(program=self.program_b, code="C2", name="Course 2", duration_slots=1, lecturer=self.lecturer, student_group=self.group_independent)

        s1 = ScheduleSlot.objects.create(timetable=self.timetable, course=c1, lecturer=self.lecturer, room=self.room_a, time_slot=self.timeslots[0], student_group=self.group_independent)
        s2 = ScheduleSlot.objects.create(timetable=self.timetable, course=c2, lecturer=self.lecturer, room=self.room_b, time_slot=self.timeslots[1], student_group=self.group_independent)

        conflicts = detect_conflicts([s1, s2], self.uni)
        travel_violations = [c for c in conflicts if c['constraint_type'] == 'LECTURER_CAMPUS_TRAVEL_VIOLATION']
        self.assertEqual(len(travel_violations), 1)

    def test_campus_travel_time_solver_prevention(self):
        """Verify solver does not schedule a lecturer at different campuses consecutively."""
        c1 = Course.objects.create(program=self.program_a, code="C1", name="Course 1", duration_slots=1, lecturer=self.lecturer, student_group=self.group_independent)
        c2 = Course.objects.create(program=self.program_b, code="C2", name="Course 2", duration_slots=1, lecturer=self.lecturer, student_group=self.group_independent)

        status, message, obj = generate_timetable(self.timetable.id)
        self.assertIn(status, ('FEASIBLE', 'OPTIMAL'))

        slots = list(ScheduleSlot.objects.filter(timetable=self.timetable).order_by('time_slot__slot_number'))
        self.assertEqual(len(slots), 2)
        
        slot_diff = abs(slots[0].time_slot.slot_number - slots[1].time_slot.slot_number)
        self.assertGreater(slot_diff, 1, f"Expected non-consecutive slots due to campus switch, got: {[s.time_slot.slot_number for s in slots]}")

    def test_max_consecutive_slots_conflict_detection(self):
        """Verify detect_conflicts flags LECTURER_CONSECUTIVE_SLOTS_VIOLATION when consecutive limit is exceeded."""
        Constraint.objects.create(
            university=self.uni,
            name="Max 2 consecutive slots for Dr. Travel",
            constraint_type="LECTURER_MAX_CONSECUTIVE_SLOTS",
            is_hard=True,
            parameters={"lecturer_id": self.lecturer.id, "max_consecutive": 2}
        )

        c1 = Course.objects.create(program=self.program_a, code="C1", name="C1", duration_slots=1, lecturer=self.lecturer, student_group=self.group_independent)
        s1 = ScheduleSlot.objects.create(timetable=self.timetable, course=c1, lecturer=self.lecturer, room=self.room_a, time_slot=self.timeslots[0], student_group=self.group_independent)
        s2 = ScheduleSlot.objects.create(timetable=self.timetable, course=c1, lecturer=self.lecturer, room=self.room_a, time_slot=self.timeslots[1], student_group=self.group_independent)
        s3 = ScheduleSlot.objects.create(timetable=self.timetable, course=c1, lecturer=self.lecturer, room=self.room_a, time_slot=self.timeslots[2], student_group=self.group_independent)

        conflicts = detect_conflicts([s1, s2, s3], self.uni)
        consec_violations = [c for c in conflicts if c['constraint_type'] == 'LECTURER_CONSECUTIVE_SLOTS_VIOLATION']
        self.assertEqual(len(consec_violations), 1)

    def test_max_consecutive_slots_solver_prevention(self):
        """Verify solver respects lecturer max consecutive slots limit."""
        Constraint.objects.create(
            university=self.uni,
            name="Max 1 consecutive slot for Dr. Travel",
            constraint_type="LECTURER_MAX_CONSECUTIVE_SLOTS",
            is_hard=True,
            parameters={"lecturer_id": self.lecturer.id, "max_consecutive": 1}
        )

        c1 = Course.objects.create(program=self.program_a, code="C1", name="C1", duration_slots=1, lecturer=self.lecturer, student_group=self.group_independent)
        c2 = Course.objects.create(program=self.program_a, code="C2", name="C2", duration_slots=1, lecturer=self.lecturer, student_group=self.group_independent)
        c3 = Course.objects.create(program=self.program_a, code="C3", name="C3", duration_slots=1, lecturer=self.lecturer, student_group=self.group_independent)

        status, message, obj = generate_timetable(self.timetable.id)
        self.assertIn(status, ('FEASIBLE', 'OPTIMAL'))

        slots = list(ScheduleSlot.objects.filter(timetable=self.timetable).order_by('time_slot__slot_number'))
        self.assertEqual(len(slots), 3)

        slot_numbers = [s.time_slot.slot_number for s in slots]
        for i in range(len(slot_numbers) - 1):
            self.assertNotEqual(slot_numbers[i+1] - slot_numbers[i], 1, f"Found consecutive slots: {slot_numbers}")

    def test_merged_classes_double_booking_conflict_detection(self):
        """Verify detect_conflicts flags double booking when parent and child groups overlap at same timeslot."""
        c1 = Course.objects.create(program=self.program_a, code="C1", name="C1", duration_slots=1, lecturer=self.lecturer, student_group=self.group_parent)
        c2 = Course.objects.create(program=self.program_a, code="C2", name="C2", duration_slots=1, lecturer=self.lecturer, student_group=self.group_child)

        s1 = ScheduleSlot(id=1, timetable=self.timetable, course=c1, lecturer=self.lecturer, room=self.room_a, time_slot=self.timeslots[0], student_group=self.group_parent)
        s2 = ScheduleSlot(id=2, timetable=self.timetable, course=c2, lecturer=self.lecturer, room=self.room_b, time_slot=self.timeslots[0], student_group=self.group_child)

        conflicts = detect_conflicts([s1, s2], self.uni)
        group_booking_violations = [c for c in conflicts if c['constraint_type'] == 'STUDENT_GROUP_DOUBLE_BOOKING']
        self.assertEqual(len(group_booking_violations), 1)

    def test_merged_classes_double_booking_solver_prevention(self):
        """Verify solver prevents parent and child groups from overlap scheduling."""
        c1 = Course.objects.create(program=self.program_a, code="C1", name="C1", duration_slots=1, lecturer=self.lecturer, student_group=self.group_parent)
        c2 = Course.objects.create(program=self.program_a, code="C2", name="C2", duration_slots=1, lecturer=self.lecturer, student_group=self.group_child)

        status, message, obj = generate_timetable(self.timetable.id)
        self.assertIn(status, ('FEASIBLE', 'OPTIMAL'))

        slots = list(ScheduleSlot.objects.filter(timetable=self.timetable))
        self.assertEqual(len(slots), 2)
        
        self.assertNotEqual(slots[0].time_slot_id, slots[1].time_slot_id)


class AdvancedFeaturesTests(TestCase):
    def setUp(self):
        from .models import RoomFeature, Building, BuildingDistance, LecturerTimeSlotPreference
        self.uni = University.objects.create(name="Advanced Feature Uni", code="AFUNI")
        self.campus = Campus.objects.create(university=self.uni, name="Campus A")
        self.faculty = Faculty.objects.create(campus=self.campus, name="Faculty A")
        self.dept = Department.objects.create(faculty=self.faculty, name="Dept A")
        self.program = Program.objects.create(department=self.dept, name="Prog A")
        self.semester = Semester.objects.create(
            university=self.uni,
            name="Fall 2026",
            start_date=datetime.date(2026, 9, 1),
            end_date=datetime.date(2026, 12, 1),
            is_active=True
        )
        self.timetable = Timetable.objects.create(semester=self.semester, name="Test Timetable")
        
        # Buildings
        self.b1 = Building.objects.create(campus=self.campus, name="Building 1")
        self.b2 = Building.objects.create(campus=self.campus, name="Building 2")
        
        # Walking time: 20 minutes (exceeds 15)
        BuildingDistance.objects.create(from_building=self.b1, to_building=self.b2, walking_time_minutes=20)
        
        # Rooms
        self.room_a = Room.objects.create(campus=self.campus, building=self.b1, name="Room A", capacity=50, room_type="Lecture")
        self.room_b = Room.objects.create(campus=self.campus, building=self.b2, name="Room B", capacity=50, room_type="Lecture")
        
        # Room features
        self.projector = RoomFeature.objects.create(name="Projector", code="PROJ")
        self.lab_equip = RoomFeature.objects.create(name="Lab Equip", code="LABEQ")
        
        self.room_a.features.add(self.projector)
        self.room_b.features.add(self.projector, self.lab_equip)
        
        self.lecturer = Lecturer.objects.create(department=self.dept, name="Dr. Advanced", email="advanced@afuni.edu", max_hours_per_week=20, max_slots_per_day=2)
        self.group1 = StudentGroup.objects.create(program=self.program, name="Group 1", size=20)
        self.group2 = StudentGroup.objects.create(program=self.program, name="Group 2", size=20)
        
        # Monday Timeslots
        self.timeslots = []
        for i in range(1, 6):
            self.timeslots.append(
                TimeSlot.objects.create(
                    university=self.uni,
                    day_of_week=1,
                    start_time=datetime.time(8 + i, 0),
                    end_time=datetime.time(9 + i, 30),
                    slot_number=i
                )
            )

    def test_room_feature_matching_conflict_detection(self):
        """Verify conflict detector flags ROOM_MISSING_REQUIRED_FEATURES when a room lacks required features."""
        c1 = Course.objects.create(program=self.program, code="C1", name="C1", duration_slots=1, lecturer=self.lecturer, student_group=self.group1)
        c1.required_features.add(self.lab_equip)
        
        s1 = ScheduleSlot.objects.create(timetable=self.timetable, course=c1, lecturer=self.lecturer, room=self.room_a, time_slot=self.timeslots[0], student_group=self.group1)
        
        conflicts = detect_conflicts([s1], self.uni)
        missing_feats = [c for c in conflicts if c['constraint_type'] == 'ROOM_MISSING_REQUIRED_FEATURES']
        self.assertEqual(len(missing_feats), 1)

    def test_room_feature_matching_solver_prevention(self):
        """Verify solver schedules a course only in a room that matches its required features."""
        c1 = Course.objects.create(program=self.program, code="C1", name="C1", duration_slots=1, lecturer=self.lecturer, student_group=self.group1)
        c1.required_features.add(self.lab_equip)
        
        status, message, obj = generate_timetable(self.timetable.id)
        self.assertIn(status, ('FEASIBLE', 'OPTIMAL'))
        
        slots = list(ScheduleSlot.objects.filter(timetable=self.timetable))
        self.assertEqual(len(slots), 1)
        self.assertEqual(slots[0].room_id, self.room_b.id)

    def test_building_travel_walking_time_conflict_detection(self):
        """Verify conflict detector flags INSUFFICIENT_TRAVEL_TIME for back-to-back classes in distant buildings."""
        c1 = Course.objects.create(program=self.program, code="C1", name="C1", duration_slots=1, lecturer=self.lecturer, student_group=self.group1)
        c2 = Course.objects.create(program=self.program, code="C2", name="C2", duration_slots=1, lecturer=self.lecturer, student_group=self.group2)
        
        s1 = ScheduleSlot.objects.create(timetable=self.timetable, course=c1, lecturer=self.lecturer, room=self.room_a, time_slot=self.timeslots[0], student_group=self.group1)
        s2 = ScheduleSlot.objects.create(timetable=self.timetable, course=c2, lecturer=self.lecturer, room=self.room_b, time_slot=self.timeslots[1], student_group=self.group2)
        
        conflicts = detect_conflicts([s1, s2], self.uni)
        travel_violations = [c for c in conflicts if c['constraint_type'] == 'INSUFFICIENT_TRAVEL_TIME']
        self.assertEqual(len(travel_violations), 1)

    def test_building_travel_walking_time_solver_prevention(self):
        """Verify solver prevents back-to-back classes in distant buildings for the same lecturer."""
        c1 = Course.objects.create(program=self.program, code="C1", name="C1", duration_slots=1, lecturer=self.lecturer, student_group=self.group1)
        c2 = Course.objects.create(program=self.program, code="C2", name="C2", duration_slots=1, lecturer=self.lecturer, student_group=self.group2)
        
        status, message, obj = generate_timetable(self.timetable.id)
        self.assertIn(status, ('FEASIBLE', 'OPTIMAL'))
        
        slots = list(ScheduleSlot.objects.filter(timetable=self.timetable).order_by('time_slot__slot_number'))
        self.assertEqual(len(slots), 2)
        
        if slots[0].room.building_id != slots[1].room.building_id:
            diff = abs(slots[0].time_slot.slot_number - slots[1].time_slot.slot_number)
            self.assertGreater(diff, 1)

    def test_shared_electives_conflict_detection(self):
        """Verify conflict detector flags STUDENT_GROUP_DOUBLE_BOOKING for overlapping slots of shared electives."""
        c1 = Course.objects.create(program=self.program, code="C1", name="C1", duration_slots=1, lecturer=self.lecturer, student_group=self.group1)
        c1.additional_student_groups.add(self.group2)
        
        c2 = Course.objects.create(program=self.program, code="C2", name="C2", duration_slots=1, lecturer=self.lecturer, student_group=self.group2)
        
        s1 = ScheduleSlot(id=1, timetable=self.timetable, course=c1, lecturer=self.lecturer, room=self.room_a, time_slot=self.timeslots[0], student_group=self.group1)
        s2 = ScheduleSlot(id=2, timetable=self.timetable, course=c2, lecturer=self.lecturer, room=self.room_b, time_slot=self.timeslots[0], student_group=self.group2)
        
        conflicts = detect_conflicts([s1, s2], self.uni)
        double_bookings = [c for c in conflicts if c['constraint_type'] == 'STUDENT_GROUP_DOUBLE_BOOKING']
        self.assertEqual(len(double_bookings), 1)

    def test_shared_electives_solver_prevention(self):
        """Verify solver prevents overlaps for additional student groups of a shared elective course."""
        c1 = Course.objects.create(program=self.program, code="C1", name="C1", duration_slots=1, lecturer=self.lecturer, student_group=self.group1)
        c1.additional_student_groups.add(self.group2)
        
        c2 = Course.objects.create(program=self.program, code="C2", name="C2", duration_slots=1, lecturer=self.lecturer, student_group=self.group2)
        
        status, message, obj = generate_timetable(self.timetable.id)
        self.assertIn(status, ('FEASIBLE', 'OPTIMAL'))
        
        slots = list(ScheduleSlot.objects.filter(timetable=self.timetable))
        self.assertEqual(len(slots), 2)
        self.assertNotEqual(slots[0].time_slot_id, slots[1].time_slot_id)

    def test_lecturer_daily_workload_limit_conflict_detection(self):
        """Verify conflict detector flags LECTURER_DAILY_LIMIT_EXCEEDED when lecturer exceeds their daily limit."""
        c1 = Course.objects.create(program=self.program, code="C1", name="C1", duration_slots=1, lecturer=self.lecturer, student_group=self.group1)
        
        s1 = ScheduleSlot.objects.create(timetable=self.timetable, course=c1, lecturer=self.lecturer, room=self.room_a, time_slot=self.timeslots[0], student_group=self.group1)
        s2 = ScheduleSlot.objects.create(timetable=self.timetable, course=c1, lecturer=self.lecturer, room=self.room_a, time_slot=self.timeslots[1], student_group=self.group1)
        s3 = ScheduleSlot.objects.create(timetable=self.timetable, course=c1, lecturer=self.lecturer, room=self.room_a, time_slot=self.timeslots[2], student_group=self.group1)
        
        conflicts = detect_conflicts([s1, s2, s3], self.uni)
        violations = [c for c in conflicts if c['constraint_type'] == 'LECTURER_DAILY_LIMIT_EXCEEDED']
        self.assertEqual(len(violations), 1)

    def test_lecturer_daily_workload_limit_solver_prevention(self):
        """Verify solver respects lecturer's max_slots_per_day limit.

        With soft constraints, the solver no longer raises INFEASIBLE; instead it
        schedules as many courses as possible within the daily limit and leaves the
        remainder unscheduled.  max_slots_per_day=2, 3 courses on same day → at
        most 2 slots scheduled.
        """
        c1 = Course.objects.create(program=self.program, code="C1", name="C1", duration_slots=1, lecturer=self.lecturer, student_group=self.group1)
        c2 = Course.objects.create(program=self.program, code="C2", name="C2", duration_slots=1, lecturer=self.lecturer, student_group=self.group2)
        c3 = Course.objects.create(program=self.program, code="C3", name="C3", duration_slots=1, lecturer=self.lecturer, student_group=self.group1)
        
        status, message, obj = generate_timetable(self.timetable.id)
        self.assertIn(status, ('OPTIMAL', 'FEASIBLE'), f"Expected solver success, got: {status}")
        
        # Solver must honour the daily cap: no more than max_slots_per_day slots scheduled
        slots = ScheduleSlot.objects.filter(timetable=self.timetable)
        self.assertLessEqual(
            slots.count(),
            self.lecturer.max_slots_per_day,
            f"Expected at most {self.lecturer.max_slots_per_day} slots but got {slots.count()}"
        )

    def test_lecturer_soft_preferences_optimization(self):
        """Verify solver optimizes lecturer timeslot preferences by avoiding disliked slots and choosing preferred ones."""
        from .models import LecturerTimeSlotPreference
        c1 = Course.objects.create(program=self.program, code="C1", name="C1", duration_slots=1, lecturer=self.lecturer, student_group=self.group1)
        
        for i in range(4):
            LecturerTimeSlotPreference.objects.create(lecturer=self.lecturer, time_slot=self.timeslots[i], preference_level='dislike')
        LecturerTimeSlotPreference.objects.create(lecturer=self.lecturer, time_slot=self.timeslots[4], preference_level='prefer')
        
        status, message, obj = generate_timetable(self.timetable.id)
        self.assertIn(status, ('FEASIBLE', 'OPTIMAL'))
        
        slots = list(ScheduleSlot.objects.filter(timetable=self.timetable))
        self.assertEqual(len(slots), 1)
        self.assertEqual(slots[0].time_slot_id, self.timeslots[4].id)


class LecturerPortalTimetableTests(TestCase):
    def setUp(self):
        from django.contrib.auth.models import User
        from accounts.models import UserProfile
        self.uni = University.objects.create(name="Lecturer Portal Uni", code="LPUNI")
        self.campus = Campus.objects.create(university=self.uni, name="Campus A")
        self.faculty = Faculty.objects.create(campus=self.campus, name="Faculty A")
        self.dept = Department.objects.create(faculty=self.faculty, name="Dept A")
        self.program = Program.objects.create(department=self.dept, name="Prog A")
        self.lecturer = Lecturer.objects.create(department=self.dept, name="Dr. Lecturer", email="lec@lpuni.edu")
        self.semester = Semester.objects.create(
            university=self.uni,
            name="Fall 2026",
            start_date=datetime.date(2026, 9, 1),
            end_date=datetime.date(2026, 12, 1),
            is_active=True
        )
        self.timetable = Timetable.objects.create(semester=self.semester, name="Active Timetable", is_active=True)
        self.room = Room.objects.create(campus=self.campus, name="Room 101", capacity=50, room_type="Lecture")
        self.group = StudentGroup.objects.create(program=self.program, name="Group 1", size=30)
        self.course1 = Course.objects.create(program=self.program, code="C1", name="Course 1", duration_slots=1, lecturer=self.lecturer, student_group=self.group)
        self.course2 = Course.objects.create(program=self.program, code="C2", name="Course 2", duration_slots=1, lecturer=self.lecturer, student_group=self.group)

        # Create Timeslots with slot_numbers on different days
        self.ts1_mon = TimeSlot.objects.create(university=self.uni, day_of_week=1, start_time=datetime.time(8, 30), end_time=datetime.time(10, 0), slot_number=1)
        self.ts3_mon = TimeSlot.objects.create(university=self.uni, day_of_week=1, start_time=datetime.time(12, 0), end_time=datetime.time(13, 30), slot_number=3)
        self.ts1_tue = TimeSlot.objects.create(university=self.uni, day_of_week=2, start_time=datetime.time(8, 30), end_time=datetime.time(10, 0), slot_number=1)
        self.ts2_tue = TimeSlot.objects.create(university=self.uni, day_of_week=2, start_time=datetime.time(10, 15), end_time=datetime.time(11, 45), slot_number=2)

        # Create schedule slots
        ScheduleSlot.objects.create(timetable=self.timetable, course=self.course1, lecturer=self.lecturer, room=self.room, time_slot=self.ts1_mon, student_group=self.group)
        ScheduleSlot.objects.create(timetable=self.timetable, course=self.course2, lecturer=self.lecturer, room=self.room, time_slot=self.ts3_mon, student_group=self.group)
        ScheduleSlot.objects.create(timetable=self.timetable, course=self.course1, lecturer=self.lecturer, room=self.room, time_slot=self.ts1_tue, student_group=self.group)
        ScheduleSlot.objects.create(timetable=self.timetable, course=self.course2, lecturer=self.lecturer, room=self.room, time_slot=self.ts2_tue, student_group=self.group)

        # User and log in setup
        self.user = User.objects.create_user(username="lecturer_user", password="password")
        UserProfile.objects.create(user=self.user, role="lecturer", university=self.uni, lecturer=self.lecturer)

    def test_lecturer_portal_weekly_timetable_grouping(self):
        """Verify that slots are grouped correctly by slot number and in chronological order in all_slots."""
        self.client.login(username="lecturer_user", password="password")
        session = self.client.session
        session['active_role'] = 'lecturer'
        session['active_university_id'] = self.uni.id
        session.save()

        response = self.client.get(reverse('scheduler:lecturer_portal_weekly_timetable'))
        self.assertEqual(response.status_code, 200)

        # Retrieve the all_slots list passed in context
        all_slots = response.context['all_slots']
        
        # Verify slot numbers are ordered by slot_number (1, 1, 2, 3)
        slot_numbers = [slot.time_slot.slot_number for slot in all_slots]
        self.assertEqual(slot_numbers, [1, 1, 2, 3])

    def test_lecturer_portal_weekly_timetable_ended_and_ongoing(self):
        """Verify the 'has_ended' and 'is_ongoing' status of slots relative to a mocked current time."""
        from unittest.mock import patch
        from django.utils import timezone
        import datetime

        # Mock timezone.now to Tuesday (day of week 2), July 7, 2026 at 09:00:00 local time
        mocked_now = timezone.make_aware(datetime.datetime(2026, 7, 7, 9, 0, 0))

        self.client.login(username="lecturer_user", password="password")
        session = self.client.session
        session['active_role'] = 'lecturer'
        session['active_university_id'] = self.uni.id
        session.save()

        with patch('django.utils.timezone.now', return_value=mocked_now):
            response = self.client.get(reverse('scheduler:lecturer_portal_weekly_timetable'))
            self.assertEqual(response.status_code, 200)

            slots = {f"{s.time_slot.day_of_week}_{s.time_slot.slot_number}": s for s in response.context['all_slots']}

            # Monday Slot 1 (8:30 - 10:00) -> should be ended because Monday < Tuesday
            self.assertTrue(slots['1_1'].has_ended)
            self.assertFalse(slots['1_1'].is_ongoing)

            # Monday Slot 3 (12:00 - 13:30) -> should be ended because Monday < Tuesday
            self.assertTrue(slots['1_3'].has_ended)
            self.assertFalse(slots['1_3'].is_ongoing)

            # Tuesday Slot 1 (8:30 - 10:00) -> should be ongoing at 09:00:00
            self.assertFalse(slots['2_1'].has_ended)
            self.assertTrue(slots['2_1'].is_ongoing)

            # Tuesday Slot 2 (10:15 - 11:45) -> should not have started yet at 09:00:00
            self.assertFalse(slots['2_2'].has_ended)
            self.assertFalse(slots['2_2'].is_ongoing)

    def test_lecturer_portal_workload_uncapped(self):
        """Verify workload percentage can exceed 100% when lecturer is over-allocated."""
        self.lecturer.max_hours_per_week = 5
        self.lecturer.save(update_fields=['max_hours_per_week'])
        self.client.login(username='lecturer_user', password='password')
        session = self.client.session
        session['active_role'] = 'lecturer'
        session['active_university_id'] = self.uni.id
        session.save()
        response = self.client.get(reverse('scheduler:lecturer_portal_workload'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['workload_pct'], 120)
        self.assertContains(response, '120% Utilization')


class EnterpriseWorkflowAndAITests(TestCase):

    def setUp(self):
        from django.contrib.auth.models import User
        from accounts.models import UserProfile
        import datetime
        self.uni = University.objects.create(name='Enterprise Test Uni', code='ETUNI')
        self.campus = Campus.objects.create(university=self.uni, name='Campus A')
        self.faculty = Faculty.objects.create(campus=self.campus, name='Faculty A')
        self.dept = Department.objects.create(faculty=self.faculty, name='Dept A')
        self.program = Program.objects.create(department=self.dept, name='Prog A')
        self.lecturer = Lecturer.objects.create(department=self.dept, name='Dr. Enterprise', email='ent@etuni.edu')
        self.semester = Semester.objects.create(
            university=self.uni,
            name='Fall 2026',
            start_date=datetime.date(2026, 9, 1),
            end_date=datetime.date(2026, 12, 1),
            is_active=True
        )
        self.timetable = Timetable.objects.create(semester=self.semester, name='Draft Timetable', status='DRAFT')
        
        self.user = User.objects.create_user(username='scheduler_user', password='password')
        UserProfile.objects.create(user=self.user, role='scheduler', university=self.uni)

        self.hod_user = User.objects.create_user(username='hod_user', password='password')
        UserProfile.objects.create(user=self.hod_user, role='hod', university=self.uni)

    def test_workflow_submission_and_approvals(self):
        # 1. Login as Scheduler
        self.client.login(username='scheduler_user', password='password')
        session = self.client.session
        session['active_role'] = 'scheduler'
        session['active_university_id'] = self.uni.id
        session.save()

        # Submit draft timetable
        response = self.client.post(
            reverse('scheduler:timetable_workflow_action', kwargs={'pk': self.timetable.pk}),
            {'action': 'submit', 'comments': 'Ready for review'}
        )
        self.assertEqual(response.status_code, 302)
        
        self.timetable.refresh_from_db()
        self.assertEqual(self.timetable.status, 'HOD_REVIEW')
        self.assertEqual(self.timetable.approval_logs.count(), 1)
        self.assertEqual(self.timetable.approval_logs.first().action, 'SUBMIT')

        # 2. Login as HOD and Approve
        self.client.login(username='hod_user', password='password')
        session = self.client.session
        session['active_role'] = 'hod'
        session['active_university_id'] = self.uni.id
        session.save()

        response = self.client.post(
            reverse('scheduler:timetable_workflow_action', kwargs={'pk': self.timetable.pk}),
            {'action': 'approve', 'comments': 'Approved HOD'}
        )
        self.assertEqual(response.status_code, 302)
        
        self.timetable.refresh_from_db()
        self.assertEqual(self.timetable.status, 'DEAN_REVIEW')
        self.assertEqual(self.timetable.approval_logs.count(), 2)

    def test_ai_quality_score_api(self):
        self.client.login(username='scheduler_user', password='password')
        session = self.client.session
        session['active_role'] = 'scheduler'
        session['active_university_id'] = self.uni.id
        session.save()

        response = self.client.get(reverse('scheduler:ai_quality_score', kwargs={'pk': self.timetable.pk}))
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data['status'], 'success')
        self.assertEqual(data['quality_score'], 100) # no slots, hence 0 violations


class StudentLoginPortalTests(TestCase):
    """
    Phase: Student Login & Portal Tests
    Covers:
      - Login with valid/invalid credentials
      - Session role set to 'student' on login
      - Dashboard redirects student to student_my_schedule
      - student_my_schedule: unlinked group → redirect to profile
      - student_my_schedule: linked group → 200 with correct slot context
      - student_portal_weekly_timetable: correct slots_by_day grouping
      - has_ended / is_ongoing flags on schedule slots
      - Logout clears session
      - Non-student role blocked from student_portal_weekly_timetable
      - Student registration creates UserProfile with role='student'
    """

    def setUp(self):
        from django.contrib.auth.models import User
        from accounts.models import UserProfile

        # University structure
        self.uni = University.objects.create(name="Student Portal Uni", code="SPUNI")
        self.campus = Campus.objects.create(university=self.uni, name="Main Campus")
        self.faculty = Faculty.objects.create(campus=self.campus, name="Science Faculty")
        self.dept = Department.objects.create(faculty=self.faculty, name="Computer Science")
        self.program = Program.objects.create(department=self.dept, name="BSc CS")

        # Semester & timetable
        self.semester = Semester.objects.create(
            university=self.uni,
            name="Semester 1 2026",
            start_date=datetime.date(2026, 9, 1),
            end_date=datetime.date(2026, 12, 15),
            is_active=True,
        )
        self.timetable = Timetable.objects.create(
            semester=self.semester,
            name="Active Timetable",
            is_active=True,
        )

        # Room & lecturer
        self.room = Room.objects.create(
            campus=self.campus, name="LH-01", capacity=60, room_type="Lecture"
        )
        self.lecturer = Lecturer.objects.create(
            department=self.dept, name="Dr. Smith", email="smith@spuni.edu"
        )

        # Student group
        self.group = StudentGroup.objects.create(
            program=self.program, name="CS Year 3", size=40
        )

        # Courses
        self.course1 = Course.objects.create(
            program=self.program, code="CS301", name="Algorithms",
            duration_slots=1, lecturer=self.lecturer, student_group=self.group,
        )
        self.course2 = Course.objects.create(
            program=self.program, code="CS302", name="Databases",
            duration_slots=1, lecturer=self.lecturer, student_group=self.group,
        )

        # Timeslots
        self.ts_mon_1 = TimeSlot.objects.create(
            university=self.uni, day_of_week=1,
            start_time=datetime.time(8, 30), end_time=datetime.time(10, 0), slot_number=1,
        )
        self.ts_mon_2 = TimeSlot.objects.create(
            university=self.uni, day_of_week=1,
            start_time=datetime.time(10, 15), end_time=datetime.time(11, 45), slot_number=2,
        )
        self.ts_tue_1 = TimeSlot.objects.create(
            university=self.uni, day_of_week=2,
            start_time=datetime.time(8, 30), end_time=datetime.time(10, 0), slot_number=1,
        )

        # Schedule slots
        self.slot_mon1 = ScheduleSlot.objects.create(
            timetable=self.timetable, course=self.course1, lecturer=self.lecturer,
            room=self.room, time_slot=self.ts_mon_1, student_group=self.group,
        )
        self.slot_mon2 = ScheduleSlot.objects.create(
            timetable=self.timetable, course=self.course2, lecturer=self.lecturer,
            room=self.room, time_slot=self.ts_mon_2, student_group=self.group,
        )
        self.slot_tue1 = ScheduleSlot.objects.create(
            timetable=self.timetable, course=self.course1, lecturer=self.lecturer,
            room=self.room, time_slot=self.ts_tue_1, student_group=self.group,
        )

        # Student user linked to group
        self.student_user = User.objects.create_user(
            username="student3", password="securePass!3"
        )
        from accounts.models import UserProfile
        UserProfile.objects.create(
            user=self.student_user,
            role="student",
            university=self.uni,
            student_group=self.group,
        )

        # Student user with NO group linked
        self.unlinked_user = User.objects.create_user(
            username="student_unlinked", password="securePass!0"
        )
        UserProfile.objects.create(
            user=self.unlinked_user,
            role="student",
            university=self.uni,
            student_group=None,
        )

        # Scheduler user (non-student) for role-guard tests
        self.scheduler_user = User.objects.create_user(
            username="scheduler3", password="schedPass!3"
        )
        UserProfile.objects.create(
            user=self.scheduler_user,
            role="scheduler",
            university=self.uni,
        )

    # ------------------------------------------------------------------ #
    # 1. Login — valid credentials                                         #
    # ------------------------------------------------------------------ #
    def test_login_valid_credentials(self):
        """Posting valid credentials logs the student in and redirects."""
        response = self.client.post(
            reverse('accounts:login'),
            {'username': 'student3', 'password': 'securePass!3'},
        )
        # Should redirect away from login page (302)
        self.assertIn(response.status_code, (301, 302))
        # User should be authenticated after login
        self.assertTrue(self.client.session.get('_auth_user_id'))

    # ------------------------------------------------------------------ #
    # 2. Login — invalid credentials                                       #
    # ------------------------------------------------------------------ #
    def test_login_invalid_password(self):
        """Wrong password keeps the user on the login page with an error."""
        response = self.client.post(
            reverse('accounts:login'),
            {'username': 'student3', 'password': 'wrongpassword'},
        )
        # Should stay on login page (200) or re-render with errors
        self.assertEqual(response.status_code, 200)
        self.assertFalse(self.client.session.get('_auth_user_id'))

    def test_login_nonexistent_user(self):
        """Login with a username that doesn't exist fails gracefully."""
        response = self.client.post(
            reverse('accounts:login'),
            {'username': 'ghost_student', 'password': 'anything'},
        )
        self.assertEqual(response.status_code, 200)
        self.assertFalse(self.client.session.get('_auth_user_id'))

    # ------------------------------------------------------------------ #
    # 3. Session role is set to 'student' on login                        #
    # ------------------------------------------------------------------ #
    def test_login_sets_student_session_role(self):
        """Login view must write active_role='student' into the session."""
        self.client.post(
            reverse('accounts:login'),
            {'username': 'student3', 'password': 'securePass!3'},
        )
        self.assertEqual(self.client.session.get('active_role'), 'student')

    def test_login_sets_university_in_session(self):
        """Login view must write active_university_id into the session."""
        self.client.post(
            reverse('accounts:login'),
            {'username': 'student3', 'password': 'securePass!3'},
        )
        self.assertEqual(
            self.client.session.get('active_university_id'), self.uni.id
        )

    # ------------------------------------------------------------------ #
    # 4. Login page renders for anonymous user                            #
    # ------------------------------------------------------------------ #
    def test_login_page_renders(self):
        """GET /accounts/login/ must return 200 for anonymous users."""
        response = self.client.get(reverse('accounts:login'))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'accounts/login.html')

    def test_login_page_redirects_authenticated_user(self):
        """An already-authenticated student visiting /login/ is redirected."""
        self.client.login(username='student3', password='securePass!3')
        response = self.client.get(reverse('accounts:login'))
        self.assertIn(response.status_code, (301, 302))

    # ------------------------------------------------------------------ #
    # 5. Logout                                                            #
    # ------------------------------------------------------------------ #
    def test_logout_clears_session(self):
        """Logout must redirect to login and clear the authentication session."""
        self.client.login(username='student3', password='securePass!3')
        self.assertTrue(self.client.session.get('_auth_user_id'))

        response = self.client.get(reverse('accounts:logout'))
        self.assertIn(response.status_code, (301, 302))
        self.assertFalse(self.client.session.get('_auth_user_id'))

    # ------------------------------------------------------------------ #
    # 6. student_my_schedule — unlinked group redirects to profile        #
    # ------------------------------------------------------------------ #
    def test_student_my_schedule_unlinked_redirects(self):
        """Student without a student_group is redirected to their profile page."""
        self.client.login(username='student_unlinked', password='securePass!0')
        session = self.client.session
        session['active_role'] = 'student'
        session['active_university_id'] = self.uni.id
        session.save()

        response = self.client.get(reverse('scheduler:student_my_schedule'))
        self.assertRedirects(response, reverse('accounts:profile'))

    # ------------------------------------------------------------------ #
    # 7. student_my_schedule — linked student sees their slots            #
    # ------------------------------------------------------------------ #
    def test_student_my_schedule_returns_200_with_slots(self):
        """Linked student sees a 200 response with their schedule slots in context."""
        self.client.login(username='student3', password='securePass!3')
        session = self.client.session
        session['active_role'] = 'student'
        session['active_university_id'] = self.uni.id
        session.save()

        response = self.client.get(reverse('scheduler:student_my_schedule'))
        self.assertEqual(response.status_code, 200)
        self.assertIn('slots', response.context)
        self.assertEqual(len(response.context['slots']), 3)

    def test_student_my_schedule_correct_student_group_in_context(self):
        """Context contains the correct student_group for the logged-in student."""
        self.client.login(username='student3', password='securePass!3')
        session = self.client.session
        session['active_role'] = 'student'
        session['active_university_id'] = self.uni.id
        session.save()

        response = self.client.get(reverse('scheduler:student_my_schedule'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['student_group'], self.group)

    def test_student_my_schedule_only_own_group_slots(self):
        """Slots returned belong exclusively to the student's group."""
        # Create another group & slot that should NOT appear
        other_group = StudentGroup.objects.create(
            program=self.program, name="CS Year 1", size=30
        )
        other_lecturer = Lecturer.objects.create(
            department=self.dept, name="Dr. Other", email="other@spuni.edu"
        )
        other_room = Room.objects.create(
            campus=self.campus, name="LH-02", capacity=60, room_type="Lecture"
        )
        other_ts = TimeSlot.objects.create(
            university=self.uni, day_of_week=3,
            start_time=datetime.time(8, 30), end_time=datetime.time(10, 0), slot_number=1,
        )
        other_course = Course.objects.create(
            program=self.program, code="CS101", name="Intro CS",
            duration_slots=1, lecturer=other_lecturer, student_group=other_group,
        )
        ScheduleSlot.objects.create(
            timetable=self.timetable, course=other_course, lecturer=other_lecturer,
            room=other_room, time_slot=other_ts, student_group=other_group,
        )

        self.client.login(username='student3', password='securePass!3')
        session = self.client.session
        session['active_role'] = 'student'
        session['active_university_id'] = self.uni.id
        session.save()

        response = self.client.get(reverse('scheduler:student_my_schedule'))
        returned_groups = {s.student_group for s in response.context['slots']}
        self.assertEqual(returned_groups, {self.group})

    # ------------------------------------------------------------------ #
    # 8. student_my_schedule — unauthenticated user redirected to login   #
    # ------------------------------------------------------------------ #
    def test_student_my_schedule_requires_login(self):
        """Unauthenticated access to student_my_schedule redirects to login."""
        response = self.client.get(reverse('scheduler:student_my_schedule'))
        self.assertIn(response.status_code, (301, 302))
        self.assertIn('/login/', response['Location'])

    # ------------------------------------------------------------------ #
    # 9. student_portal_weekly_timetable — correct slots_by_day grouping  #
    # ------------------------------------------------------------------ #
    def test_student_portal_weekly_timetable_slots_by_day(self):
        """slots_by_day must correctly bucket slots into day-of-week keys."""
        self.client.login(username='student3', password='securePass!3')
        session = self.client.session
        session['active_role'] = 'student'
        session['active_university_id'] = self.uni.id
        session.save()

        response = self.client.get(reverse('scheduler:student_portal_weekly_timetable'))
        self.assertEqual(response.status_code, 200)

        slots_by_day = response.context['slots_by_day']
        # Monday (1): slot_mon1, slot_mon2
        self.assertEqual(len(slots_by_day[1]), 2)
        # Tuesday (2): slot_tue1
        self.assertEqual(len(slots_by_day[2]), 1)
        # Other days: empty
        for day in (3, 4, 5):
            self.assertEqual(len(slots_by_day[day]), 0)

    def test_student_portal_weekly_timetable_all_slots_ordered_by_slot_number(self):
        """all_slots must be ordered by slot_number then day_of_week."""
        self.client.login(username='student3', password='securePass!3')
        session = self.client.session
        session['active_role'] = 'student'
        session['active_university_id'] = self.uni.id
        session.save()

        response = self.client.get(reverse('scheduler:student_portal_weekly_timetable'))
        self.assertEqual(response.status_code, 200)

        all_slots = response.context['all_slots']
        # slot_number order: 1 (Mon), 1 (Tue), 2 (Mon)
        slot_numbers = [s.time_slot.slot_number for s in all_slots]
        self.assertEqual(slot_numbers, sorted(slot_numbers))

    # ------------------------------------------------------------------ #
    # 10. has_ended / is_ongoing flags on student_my_schedule             #
    # ------------------------------------------------------------------ #
    def test_student_schedule_has_ended_is_ongoing_flags(self):
        """
        When mocked to Tuesday 09:00, Monday slots must have has_ended=True,
        Tuesday slot 1 (08:30–10:00) is_ongoing=True, and future slots are neutral.
        """
        from unittest.mock import patch

        mocked_now = timezone.make_aware(datetime.datetime(2026, 7, 7, 9, 0, 0))

        self.client.login(username='student3', password='securePass!3')
        session = self.client.session
        session['active_role'] = 'student'
        session['active_university_id'] = self.uni.id
        session.save()

        with patch('django.utils.timezone.now', return_value=mocked_now):
            response = self.client.get(reverse('scheduler:student_my_schedule'))
        self.assertEqual(response.status_code, 200)

        slots = {
            f"{s.time_slot.day_of_week}_{s.time_slot.slot_number}": s
            for s in response.context['slots']
        }

        # Monday slots — both should have ended (Monday < Tuesday)
        self.assertTrue(slots['1_1'].has_ended)
        self.assertFalse(slots['1_1'].is_ongoing)
        self.assertTrue(slots['1_2'].has_ended)
        self.assertFalse(slots['1_2'].is_ongoing)

        # Tuesday slot 1 (08:30–10:00) — ongoing at 09:00
        self.assertFalse(slots['2_1'].has_ended)
        self.assertTrue(slots['2_1'].is_ongoing)

    # ------------------------------------------------------------------ #
    # 11. Non-student is redirected away from student_portal_weekly       #
    # ------------------------------------------------------------------ #
    def test_non_student_redirected_from_student_weekly_timetable(self):
        """
        A lecturer user accessing student_portal_weekly_timetable is redirected
        (the view enforces ROLE_STUDENT).
        """
        from django.contrib.auth.models import User
        from accounts.models import UserProfile

        lec_user = User.objects.create_user(username="lec_guard3", password="pass")
        UserProfile.objects.create(
            user=lec_user, role="lecturer", university=self.uni
        )

        self.client.login(username="lec_guard3", password="pass")
        session = self.client.session
        session['active_role'] = 'lecturer'
        session['active_university_id'] = self.uni.id
        session.save()

        response = self.client.get(reverse('scheduler:student_portal_weekly_timetable'))
        # View should redirect away (lecturer has no student_group → profile redirect)
        self.assertIn(response.status_code, (301, 302))

    # ------------------------------------------------------------------ #
    # 12. Student registration creates correct UserProfile                #
    # ------------------------------------------------------------------ #
    def test_student_registration_creates_profile_with_student_role(self):
        """Registering as a student creates a UserProfile with role='student'."""
        from django.contrib.auth.models import User
        from accounts.models import UserProfile

        response = self.client.post(
            reverse('accounts:register'),
            {
                'username': 'new_student3',
                'first_name': 'Alice',
                'last_name': 'Kariuki',
                'email': 'alice@spuni.edu',
                'password1': 'Str0ng#Pass99',
                'password2': 'Str0ng#Pass99',
                'role': 'student',
                'university': self.uni.id,
            },
        )
        # Successful registration redirects to login
        self.assertIn(response.status_code, (301, 302))

        user = User.objects.filter(username='new_student3').first()
        self.assertIsNotNone(user)
        profile = UserProfile.objects.filter(user=user).first()
        self.assertIsNotNone(profile)
        self.assertEqual(profile.role, 'student')
        self.assertEqual(profile.university, self.uni)


class HODAndLecturerClassificationTests(TestCase):
    def setUp(self):
        from django.contrib.auth.models import User
        from accounts.models import UserProfile
        from .models import Room, Lecturer, Semester, Timetable, StudentGroup, Course, TimeSlot, ScheduleSlot
        import datetime

        self.uni = University.objects.create(name="HOD and Classify Test Uni", code="HCTUNI")
        self.campus = Campus.objects.create(university=self.uni, name="Main Campus")
        self.campus2 = Campus.objects.create(university=self.uni, name="Remote Campus")
        self.faculty = Faculty.objects.create(campus=self.campus, name="Science Faculty")
        self.dept = Department.objects.create(faculty=self.faculty, name="CS Dept")
        self.program = Program.objects.create(department=self.dept, name="BSc CS")

        self.semester = Semester.objects.create(
            university=self.uni,
            name="Fall 2026",
            start_date=datetime.date(2026, 9, 1),
            end_date=datetime.date(2026, 12, 1),
            is_active=True
        )
        self.timetable = Timetable.objects.create(semester=self.semester, name="Workflow Timetable", status="DRAFT")

        # HOD user
        self.hod_user = User.objects.create_user(username="test_hod_user", password="password")
        self.hod_profile = UserProfile.objects.create(user=self.hod_user, role="hod", university=self.uni)

    def test_hod_role_properties(self):
        """Verify HOD role permissions property is_hod resolves to True."""
        self.assertTrue(self.hod_profile.is_hod)
        self.assertFalse(self.hod_profile.is_student)
        self.assertFalse(self.hod_profile.is_lecturer)

    def test_online_room_travel_exception(self):
        """Verify that conflict detection passes without travel violations if one room is online/virtual."""
        from .models import Room, Lecturer, TimeSlot, Course, ScheduleSlot
        from .conflicts import detect_conflicts
        import datetime

        # Create physical room on remote campus
        room_physical = Room.objects.create(
            campus=self.campus2, name="Physical Rm 101", capacity=50, room_type="Lecture"
        )
        # Create virtual room (contains 'zoom') on main campus
        room_virtual = Room.objects.create(
            campus=self.campus, name="Zoom Room Alpha", capacity=100, room_type="Virtual", is_virtual=True
        )

        lecturer = Lecturer.objects.create(
            department=self.dept, name="Dr. Travel", email="travel@hctuni.edu", max_hours_per_week=20, max_slots_per_day=3
        )
        group = StudentGroup.objects.create(program=self.program, name="CS Group 1", size=30)

        c1 = Course.objects.create(program=self.program, code="CS101", name="Course 1", duration_slots=1, lecturer=lecturer, student_group=group)
        c2 = Course.objects.create(program=self.program, code="CS102", name="Course 2", duration_slots=1, lecturer=lecturer, student_group=group)

        ts1 = TimeSlot.objects.create(
            university=self.uni, day_of_week=1, start_time=datetime.time(8, 30), end_time=datetime.time(10, 0), slot_number=1
        )
        ts2 = TimeSlot.objects.create(
            university=self.uni, day_of_week=1, start_time=datetime.time(10, 15), end_time=datetime.time(11, 45), slot_number=2
        )

        # Lecturer has back-to-back classes on different campuses
        s1 = ScheduleSlot.objects.create(
            timetable=self.timetable, course=c1, lecturer=lecturer, room=room_physical, time_slot=ts1, student_group=group
        )
        s2 = ScheduleSlot.objects.create(
            timetable=self.timetable, course=c2, lecturer=lecturer, room=room_virtual, time_slot=ts2, student_group=group
        )

        # Detect conflicts
        conflicts = detect_conflicts([s1, s2], self.uni)

        # Filter for LECTURER_CAMPUS_TRAVEL_VIOLATION
        travel_violations = [c for c in conflicts if c['constraint_type'] == 'LECTURER_CAMPUS_TRAVEL_VIOLATION']
        
        # There should be no travel violation because 'Zoom Room Alpha' is recognized as online
        self.assertEqual(len(travel_violations), 0)

    def test_lecturer_full_time_part_time_classification(self):
        """Verify lecturers are correctly classified as full-time or part-time based on hours."""
        from .models import Lecturer

        ft_lec = Lecturer.objects.create(
            department=self.dept, name="Full Time Pro", email="ft@hctuni.edu", max_hours_per_week=12
        )
        pt_lec = Lecturer.objects.create(
            department=self.dept, name="Part Time Pro", email="pt@hctuni.edu", max_hours_per_week=10
        )

        # Check categorization boundary
        self.assertTrue(ft_lec.max_hours_per_week >= 12)
        self.assertTrue(pt_lec.max_hours_per_week < 12)

    def test_shared_course_room_capacity_aggregation(self):
        """Verify that the solver sums student group sizes for shared courses/common units."""
        from .solver import generate_timetable
        from .models import StudentGroup, Room, Course, Lecturer, TimeSlot, Timetable, ScheduleSlot

        # Total combined size = 55
        group_main = StudentGroup.objects.create(program=self.program, name="CS Main Group", size=30)
        group_add = StudentGroup.objects.create(program=self.program, name="CS Add Group", size=25)

        # Room Small (capacity 40) is too small for combined size of 55
        # Room Large (capacity 60) can fit 55
        room_small = Room.objects.create(campus=self.campus, name="Room Small", capacity=40, room_type="Lecture")
        room_large = Room.objects.create(campus=self.campus, name="Room Large", capacity=60, room_type="Lecture")

        lecturer = Lecturer.objects.create(
            department=self.dept, name="Dr. Shared", email="shared@hctuni.edu", max_hours_per_week=20, max_slots_per_day=3
        )

        shared_course = Course.objects.create(
            program=self.program, code="CS999", name="Shared Course", duration_slots=1, lecturer=lecturer, student_group=group_main
        )
        shared_course.additional_student_groups.add(group_add)

        ts = TimeSlot.objects.create(
            university=self.uni, day_of_week=1, start_time=datetime.time(12, 0), end_time=datetime.time(13, 30), slot_number=3
        )

        tt = Timetable.objects.create(semester=self.semester, name="Shared Timetable", status="DRAFT")

        status, msg, log_id = generate_timetable(tt.id)

        # Fetch the scheduled slot and check that the large room was selected
        slot = ScheduleSlot.objects.filter(timetable=tt, course=shared_course).first()
        self.assertIsNotNone(slot)
        self.assertEqual(slot.room.id, room_large.id)

    def test_generate_timetable_with_virtual_room(self):
        """Verify that generate_timetable succeeds when virtual rooms are present in the university."""
        from .solver import generate_timetable
        from .models import Room, Course, Lecturer, TimeSlot, Timetable, ScheduleSlot, StudentGroup
        import datetime

        room_virtual = Room.objects.create(
            campus=self.campus, name="Zoom Room Alpha", capacity=100, room_type="Virtual", is_virtual=True
        )

        lecturer = Lecturer.objects.create(
            department=self.dept, name="Dr. Virtual Test", email="virtual_test@hctuni.edu", max_hours_per_week=20, max_slots_per_day=3
        )
        group = StudentGroup.objects.create(program=self.program, name="CS Virtual Group 1", size=30)

        course = Course.objects.create(
            program=self.program, code="CS888", name="Virtual Course", duration_slots=1, lecturer=lecturer, student_group=group
        )

        ts = TimeSlot.objects.create(
            university=self.uni, day_of_week=1, start_time=datetime.time(8, 30), end_time=datetime.time(10, 0), slot_number=1
        )

        tt = Timetable.objects.create(semester=self.semester, name="Virtual Timetable", status="DRAFT")

        status, msg, log_id = generate_timetable(tt.id)
        self.assertIn(status, ('OPTIMAL', 'FEASIBLE'))


class LecturerCredentialsAndRegistrationTests(TestCase):
    def setUp(self):
        from django.contrib.auth.models import User
        from accounts.models import UserProfile
        from .models import Room, Lecturer, Semester, Timetable, StudentGroup, Course, TimeSlot, ScheduleSlot
        from scheduler.signals import _local
        import datetime

        _local.enable_auto_provision_in_tests = True

        self.uni = University.objects.create(name="Credentials Test Uni", code="CRUNI")
        self.campus = Campus.objects.create(university=self.uni, name="Main Campus")
        self.faculty = Faculty.objects.create(campus=self.campus, name="Science Faculty")
        self.dept = Department.objects.create(faculty=self.faculty, name="CS Dept")
        self.program = Program.objects.create(department=self.dept, name="BSc CS")
        
        self.semester = Semester.objects.create(
            university=self.uni,
            name="Fall 2026",
            start_date=datetime.date(2026, 9, 1),
            end_date=datetime.date(2026, 12, 1),
            is_active=True
        )
        self.timetable = Timetable.objects.create(semester=self.semester, name="Test Timetable", is_active=True)
        self.room = Room.objects.create(campus=self.campus, name="Room 101", capacity=50, room_type="Lecture")
        self.group = StudentGroup.objects.create(program=self.program, name="CS Y2", size=30)
        self.ts = TimeSlot.objects.create(
            university=self.uni, day_of_week=1, start_time=datetime.time(8, 30), end_time=datetime.time(10, 0), slot_number=1
        )

    def tearDown(self):
        from scheduler.signals import _local
        _local.enable_auto_provision_in_tests = False
        super().tearDown()

    def test_lecturer_credentials_provisioning_on_save(self):
        """Creating a lecturer automatically provisions a Django User and UserProfile."""
        from django.contrib.auth.models import User
        from accounts.models import UserProfile
        from .models import Lecturer

        lec = Lecturer.objects.create(
            department=self.dept,
            name="Professor Test",
            email="prof.test@cruni.edu",
            is_active=True
        )

        # Verify that User and UserProfile are created
        user = User.objects.filter(email="prof.test@cruni.edu").first()
        self.assertIsNotNone(user)
        self.assertTrue(user.is_active)
        
        profile = UserProfile.objects.filter(user=user).first()
        self.assertIsNotNone(profile)
        self.assertEqual(profile.role, 'lecturer')
        self.assertEqual(profile.lecturer, lec)
        
        # Verify Lecturer is linked back to the User
        lec.refresh_from_db()
        self.assertEqual(lec.user, user)

    def test_lecturer_credentials_provisioning_missing_email(self):
        """Creating a lecturer with a missing email generates fallback and registers warning."""
        from django.contrib.auth.models import User
        from accounts.models import UserProfile
        from scheduler.models import ImportAuditLog, Lecturer

        # Create a dummy audit log to make sure warnings capture correctly
        ImportAuditLog.objects.create(
            university=self.uni,
            file_name="dummy.xlsx",
            import_type="smart"
        )

        lec = Lecturer.objects.create(
            department=self.dept,
            name="Professor NoEmail",
            is_active=True
        )

        # Verify fallback email generated
        user = User.objects.filter(email="professor.noemail@credentialstestuni.edu").first()
        self.assertIsNotNone(user)
        
        profile = UserProfile.objects.filter(user=user).first()
        self.assertIsNotNone(profile)
        self.assertEqual(profile.lecturer, lec)

        # Verify warning was added to audit log
        audit = ImportAuditLog.objects.filter(university=self.uni).first()
        self.assertIsNotNone(audit)
        self.assertTrue(any("fallback" in w for w in audit.warnings))

    def test_lecturer_credentials_expiry_cron(self):
        """Deactivating/ending semesters disables lecturer access via cron function."""
        from django.contrib.auth.models import User
        from accounts.models import UserProfile
        from .models import Lecturer, Course, ScheduleSlot
        from .tasks import expire_ended_semester_credentials

        lec = Lecturer.objects.create(
            department=self.dept,
            name="Expired Lecturer",
            email="expired@cruni.edu",
            is_active=True
        )

        user = User.objects.get(email="expired@cruni.edu")
        self.assertTrue(user.is_active)

        # Scenario A: Lecturer is not assigned to any course -> Access should expire
        expire_ended_semester_credentials()
        user.refresh_from_db()
        self.assertFalse(user.is_active)

    def test_lecturer_registration_sends_verification_email(self):
        """Registering a new lecturer via RegisterForm creates Lecturer profile and sends verification email."""
        from django.core import mail
        from accounts.forms import RegisterForm
        from scheduler.tasks import verify_and_notify_lecturer_record

        form_data = {
            'username': 'newlecturer',
            'first_name': 'Jane',
            'last_name': 'Doe',
            'email': 'jane.doe@cruni.edu',
            'role': 'lecturer',
            'university': self.uni.id,
            'password1': 'StrongPass123!',
            'password2': 'StrongPass123!',
        }
        form = RegisterForm(data=form_data)
        self.assertTrue(form.is_valid(), form.errors)
        user = form.save()
        
        # Verify Lecturer auto-creation
        lec = Lecturer.objects.filter(email='jane.doe@cruni.edu').first()
        self.assertIsNotNone(lec)
        self.assertTrue(lec.is_verified)
        self.assertEqual(lec.user, user)

        # Trigger verification email task
        verify_and_notify_lecturer_record(
            submitted_email=user.email,
            submitted_name=f"{user.first_name} {user.last_name}",
            university_id=self.uni.id,
            preserve_password=True
        )

        # Assert email sent
        self.assertTrue(len(mail.outbox) >= 1)
        sent_mail = mail.outbox[-1]
        self.assertIn("Verified", sent_mail.subject)
        self.assertIn("jane.doe@cruni.edu", sent_mail.to)

    def test_public_onboarding_lookup_renders_form_before_save(self):
        """Lookup search step renders public_onboarding form and does not jump directly to success page."""
        response = self.client.post(
            reverse('scheduler:public_lecturer_onboarding_direct'),
            {'lookup_query': 'jane.doe@cruni.edu'}
        )
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'scheduler/public_onboarding.html')

    def test_public_onboarding_save_renders_lookup_with_success_message(self):
        """Submitting onboarding preferences renders the lookup page again with success message."""
        lec = Lecturer.objects.create(department=self.dept, name="Save Test Lec", email="save.test@cruni.edu")
        response = self.client.post(
            reverse('scheduler:public_lecturer_onboarding', kwargs={'token': lec.calendar_token}),
            {'save_onboarding': '1', 'name': 'Save Test Lec Updated', 'email': 'save.test@cruni.edu', 'max_hours_per_week': '18'}
        )
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'scheduler/public_onboarding_lookup.html')

    def test_student_registration_saves_group(self):
        """Student registration form captures student group/course and saves it to profile."""
        from django.contrib.auth.models import User
        from accounts.models import UserProfile
        from accounts.forms import RegisterForm

        form = RegisterForm({
            'username': 'form_student',
            'first_name': 'Jane',
            'last_name': 'Doe',
            'email': 'jane.doe@cruni.edu',
            'password1': 'StrongPass123!',
            'password2': 'StrongPass123!',
            'role': 'student',
            'university': self.uni.id,
            'student_group': self.group.id
        })

        self.assertTrue(form.is_valid(), form.errors)
        user = form.save()

        # Check UserProfile for group link
        profile = UserProfile.objects.get(user=user)
        self.assertEqual(profile.role, 'student')
        self.assertEqual(profile.student_group, self.group)



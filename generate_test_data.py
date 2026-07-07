import openpyxl
import random

def generate_rooms():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rooms"
    
    # Headers
    headers = ["name", "capacity", "room_type", "campus_name"]
    ws.append(headers)
    
    room_types = ["Lecture", "Lab", "Lecture Hall"]
    campuses = ["Main Campus", "North Campus", "East Campus", "Science Park"]
    
    for i in range(1, 2001):
        name = f"Room {100 + i // 4}-{chr(65 + i % 4)}"
        capacity = random.choice([30, 40, 50, 60, 80, 100, 120, 150])
        room_type = random.choice(room_types)
        campus_name = random.choice(campuses)
        ws.append([name, capacity, room_type, campus_name])
        
    wb.save("rooms_import_2000.xlsx")
    print("Generated rooms_import_2000.xlsx")

def generate_lecturers():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lecturers"
    
    # Headers
    headers = ["name", "email", "department_name", "max_hours"]
    ws.append(headers)
    
    first_names = ["James", "Mary", "John", "Patricia", "Robert", "Jennifer", "Michael", "Elizabeth", "William", "Linda", "David", "Barbara", "Richard", "Susan", "Joseph", "Jessica", "Thomas", "Sarah", "Charles", "Karen"]
    last_names = ["Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia", "Miller", "Davis", "Rodriguez", "Martinez", "Hernandez", "Lopez", "Gonzalez", "Wilson", "Anderson", "Thomas", "Taylor", "Moore", "Jackson", "Martin"]
    departments = ["Computer Science", "Mathematics", "Physics", "Chemistry", "Electrical Engineering", "Mechanical Engineering", "Business Administration", "Economics", "Literature", "History"]
    
    emails_used = set()
    
    for i in range(1, 2001):
        # Ensure unique email
        while True:
            first = random.choice(first_names)
            last = random.choice(last_names)
            email = f"{first.lower()}.{last.lower()}{i}@university.edu"
            if email not in emails_used:
                emails_used.add(email)
                break
        
        name = f"Dr. {first} {last}"
        dept = random.choice(departments)
        max_hours = random.choice([12, 16, 20, 24])
        ws.append([name, email, dept, max_hours])
        
    wb.save("lecturers_import_2000.xlsx")
    print("Generated lecturers_import_2000.xlsx")
    return list(emails_used)

def generate_student_groups():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Groups"
    
    # Headers
    headers = ["name", "size", "program_name"]
    ws.append(headers)
    
    programs = ["BSc Computer Science", "BSc Mathematics", "BSc Physics", "BSc Chemistry", "BSc Electrical Engineering", "BSc Mechanical Engineering", "Bachelor of Business Admin", "BSc Economics", "BA Literature", "BA History"]
    
    group_names = []
    for i in range(1, 2001):
        prog = random.choice(programs)
        prog_short = "".join([w[0] for w in prog.split() if w[0].isupper()])
        year = random.choice([1, 2, 3, 4])
        sec = chr(65 + (i % 3))
        name = f"{prog_short} Yr{year} Sec {sec} - G{i}"
        size = random.choice([25, 30, 35, 40, 45, 50, 60])
        ws.append([name, size, prog])
        group_names.append((name, prog))
        
    wb.save("student_groups_import_2000.xlsx")
    print("Generated student_groups_import_2000.xlsx")
    return group_names

def generate_courses(emails, groups):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Courses"
    
    # Headers
    headers = ["code", "name", "duration_slots", "required_room_type", "lecturer_email", "student_group_name", "program_name"]
    ws.append(headers)
    
    course_subjects = {
        "CS": ["Intro to Programming", "Data Structures", "Algorithms", "Database Systems", "Software Engineering", "Operating Systems", "Computer Networks", "Artificial Intelligence", "Machine Learning", "Web Development"],
        "MATH": ["Calculus I", "Calculus II", "Linear Algebra", "Discrete Mathematics", "Probability and Statistics", "Differential Equations", "Numerical Analysis", "Abstract Algebra", "Real Analysis", "Complex Variables"],
        "PHYS": ["Mechanics", "Electromagnetism", "Thermodynamics", "Quantum Physics", "Optics", "Nuclear Physics", "Solid State Physics", "Astrophysics", "Analytical Mechanics", "Statistical Physics"],
        "CHEM": ["General Chemistry", "Organic Chemistry", "Inorganic Chemistry", "Physical Chemistry", "Analytical Chemistry", "Biochemistry", "Environmental Chemistry", "Polymer Chemistry", "Spectroscopy", "Quantum Chemistry"],
        "EE": ["Circuit Analysis", "Digital Systems", "Signals and Systems", "Electromagnetics", "Microprocessors", "Control Systems", "Power Systems", "Communication Systems", "VLSI Design", "Embedded Systems"],
        "ME": ["Statics", "Dynamics", "Strength of Materials", "Fluid Mechanics", "Heat Transfer", "Thermodynamics", "Machine Design", "Manufacturing Processes", "Control Systems", "Vibrations"],
        "BUS": ["Principles of Management", "Marketing", "Financial Accounting", "Organizational Behavior", "Human Resource Management", "Strategic Management", "Operations Management", "Corporate Finance", "Business Law", "International Business"],
        "ECON": ["Microeconomics", "Macroeconomics", "Econometrics", "Game Theory", "International Trade", "Public Finance", "Labor Economics", "Development Economics", "Monetary Economics", "Financial Economics"],
        "LIT": ["English Literature", "World Literature", "Creative Writing", "Literary Theory", "Shakespeare", "Modern Poetry", "American Literature", "Comparative Literature", "Drama Studies", "Classical Mythology"],
        "HIST": ["World History", "European History", "Ancient Civilizations", "Modern History", "History of Science", "Political History", "Cultural History", "Military History", "Economic History", "Historiography"]
    }
    
    keys = list(course_subjects.keys())
    
    for i in range(1, 2001):
        subj = random.choice(keys)
        title = random.choice(course_subjects[subj])
        code = f"{subj}{100 + i}"
        name = f"{title} {i}"
        duration = random.choice([1, 2, 2, 3]) # Bias towards 2 slots
        room_type = "Lab" if subj in ["CS", "PHYS", "CHEM", "EE", "ME"] and random.random() > 0.4 else "Lecture"
        
        # Link to random lecturer and student group
        lec_email = random.choice(emails)
        group_name, prog_name = random.choice(groups)
        
        ws.append([code, name, duration, room_type, lec_email, group_name, prog_name])
        
    wb.save("courses_import_2000.xlsx")
    print("Generated courses_import_2000.xlsx")

if __name__ == "__main__":
    emails = generate_lecturers()
    groups = generate_student_groups()
    generate_rooms()
    generate_courses(emails, groups)
    print("All 4 Excel sheets with 2000 rows each generated successfully!")
